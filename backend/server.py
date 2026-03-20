from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, WebSocket, WebSocketDisconnect, Query, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any, Literal
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import asyncio
import json
from io import BytesIO
from openpyxl import load_workbook, Workbook
from xlsxwriter import Workbook as XlsxWriterWorkbook
import struct
from enum import Enum

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI(title="Tag Mapping Communication Gateway", version="1.0.0")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'comgate-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

security = HTTPBearer()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== ENUMS ====================
class UserRole(str, Enum):
    VIEWER = "viewer"
    ENGINEER = "engineer"
    OPERATOR = "operator"
    ADMIN = "admin"

class ProtocolType(str, Enum):
    TCP = "tcp"
    UDP = "udp"
    RTU = "rtu"

class ObjectType(str, Enum):
    COIL = "coil"
    DISCRETE_INPUT = "discrete_input"
    INPUT_REGISTER = "input_register"
    HOLDING_REGISTER = "holding_register"

class DataType(str, Enum):
    BOOL = "bool"
    INT16 = "int16"
    UINT16 = "uint16"
    INT32 = "int32"
    UINT32 = "uint32"
    FLOAT32 = "float32"
    FLOAT64 = "float64"
    STRING = "string"

class TagPermission(str, Enum):
    READ = "R"
    WRITE = "W"
    READ_WRITE = "RW"

class EndianType(str, Enum):
    ABCD = "ABCD"
    CDAB = "CDAB"
    BADC = "BADC"
    DCBA = "DCBA"

class TagQuality(str, Enum):
    GOOD = "good"
    BAD = "bad"
    UNCERTAIN = "uncertain"

class TrafficStatus(str, Enum):
    OK = "ok"
    ERROR = "error"
    TIMEOUT = "timeout"

# ==================== MODELS ====================
class UserBase(BaseModel):
    email: EmailStr
    username: str
    role: UserRole = UserRole.VIEWER

class UserCreate(UserBase):
    password: str

class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_active: bool = True

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: User

class ProjectBase(BaseModel):
    name: str
    description: Optional[str] = ""
    default_timeout_ms: int = 3000
    default_retries: int = 3

class Project(ProjectBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str = ""

class DeviceBase(BaseModel):
    name: str
    protocol: ProtocolType
    group: Optional[str] = ""
    is_enabled: bool = True
    # TCP/UDP settings
    ip_address: Optional[str] = None
    port: Optional[int] = None
    # RTU settings
    com_port: Optional[str] = None
    baud_rate: Optional[int] = 9600
    parity: Optional[str] = "N"
    data_bits: Optional[int] = 8
    stop_bits: Optional[int] = 1
    # Common
    unit_id: int = 1
    timeout_ms: int = 3000
    retries: int = 3
    max_block_size: int = 120
    default_endian: EndianType = EndianType.ABCD

class Device(DeviceBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    status: str = "offline"
    last_poll: Optional[datetime] = None
    error_count: int = 0
    success_count: int = 0

class TagBase(BaseModel):
    name: str
    device_id: str
    object_type: ObjectType
    address: int
    bit: Optional[int] = None
    data_type: DataType = DataType.UINT16
    permission: TagPermission = TagPermission.READ
    scale: float = 1.0
    offset: float = 0.0
    unit: Optional[str] = ""
    poll_ms: int = 1000
    min_value: Optional[float] = None
    max_value: Optional[float] = None
    endian: Optional[EndianType] = None
    description: Optional[str] = ""
    alarm_enable: bool = False
    deadband: Optional[float] = None
    write_confirm: bool = False
    is_spare: bool = False

class Tag(TagBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    current_value: Optional[Any] = None
    quality: TagQuality = TagQuality.UNCERTAIN
    last_update: Optional[datetime] = None
    error_message: Optional[str] = None

class TagValue(BaseModel):
    tag_id: str
    value: Any
    quality: TagQuality
    timestamp: datetime

class TrafficLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    device_id: str
    device_name: str
    protocol: ProtocolType
    function_code: int
    request_summary: str
    response_summary: Optional[str] = None
    round_trip_ms: Optional[float] = None
    status: TrafficStatus
    error_details: Optional[str] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    username: str
    action: str
    details: Dict[str, Any] = {}
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ImportReport(BaseModel):
    total_rows: int = 0
    success_count: int = 0
    warning_count: int = 0
    error_count: int = 0
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []

class WriteRequest(BaseModel):
    tag_id: str
    value: Any

class BatchWriteRequest(BaseModel):
    writes: List[WriteRequest]

class SimulatorConfig(BaseModel):
    protocol: ProtocolType
    port: int = 5020
    unit_id: int = 1
    num_coils: int = 100
    num_discrete_inputs: int = 100
    num_input_registers: int = 100
    num_holding_registers: int = 100

class SimulatorSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    config: SimulatorConfig
    is_running: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ==================== HELPER FUNCTIONS ====================
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user: User) -> str:
    payload = {
        "sub": user.id,
        "email": user.email,
        "role": user.role.value,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> User:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user_doc:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user_doc)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(allowed_roles: List[UserRole]):
    async def role_checker(user: User = Depends(get_current_user)):
        if user.role not in allowed_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return role_checker

async def log_audit(user: User, action: str, details: Dict[str, Any] = {}):
    audit = AuditLog(
        user_id=user.id,
        username=user.username,
        action=action,
        details=details
    )
    doc = audit.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    await db.audit_logs.insert_one(doc)

# ==================== AUTH ENDPOINTS ====================
@api_router.post("/auth/register", response_model=User)
async def register(user_create: UserCreate):
    existing = await db.users.find_one({"email": user_create.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(**user_create.model_dump(exclude={"password"}))
    doc = user.model_dump()
    doc['password_hash'] = hash_password(user_create.password)
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    return user

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user_doc or not verify_password(credentials.password, user_doc.get('password_hash', '')):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if isinstance(user_doc.get('created_at'), str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    
    user = User(**{k: v for k, v in user_doc.items() if k != 'password_hash'})
    token = create_token(user)
    return TokenResponse(access_token=token, user=user)

@api_router.get("/auth/me", response_model=User)
async def get_me(user: User = Depends(get_current_user)):
    return user

# ==================== USER MANAGEMENT ====================
@api_router.get("/users", response_model=List[User])
async def list_users(user: User = Depends(require_role([UserRole.ADMIN]))):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    for u in users:
        if isinstance(u.get('created_at'), str):
            u['created_at'] = datetime.fromisoformat(u['created_at'])
    return users

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, role: UserRole, admin: User = Depends(require_role([UserRole.ADMIN]))):
    result = await db.users.update_one({"id": user_id}, {"$set": {"role": role.value}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    await log_audit(admin, "user_role_updated", {"user_id": user_id, "new_role": role.value})
    return {"message": "Role updated"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: User = Depends(require_role([UserRole.ADMIN]))):
    if admin.id == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    await log_audit(admin, "user_deleted", {"user_id": user_id})
    return {"message": "User deleted"}

# ==================== PROJECT ENDPOINTS ====================
@api_router.post("/projects", response_model=Project)
async def create_project(project: ProjectBase, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    proj = Project(**project.model_dump(), created_by=user.id)
    doc = proj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.projects.insert_one(doc)
    await log_audit(user, "project_created", {"project_id": proj.id, "name": proj.name})
    return proj

@api_router.get("/projects", response_model=List[Project])
async def list_projects(user: User = Depends(get_current_user)):
    projects = await db.projects.find({}, {"_id": 0}).to_list(100)
    for p in projects:
        for field in ['created_at', 'updated_at']:
            if isinstance(p.get(field), str):
                p[field] = datetime.fromisoformat(p[field])
    return projects

@api_router.get("/projects/{project_id}", response_model=Project)
async def get_project(project_id: str, user: User = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    for field in ['created_at', 'updated_at']:
        if isinstance(project.get(field), str):
            project[field] = datetime.fromisoformat(project[field])
    return project

@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, update: ProjectBase, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    update_doc = update.model_dump()
    update_doc['updated_at'] = datetime.now(timezone.utc).isoformat()
    result = await db.projects.update_one({"id": project_id}, {"$set": update_doc})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    await log_audit(user, "project_updated", {"project_id": project_id})
    return await get_project(project_id, user)

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, user: User = Depends(require_role([UserRole.ADMIN]))):
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    await db.devices.delete_many({"project_id": project_id})
    await db.tags.delete_many({"project_id": project_id})
    await db.traffic_logs.delete_many({"project_id": project_id})
    await log_audit(user, "project_deleted", {"project_id": project_id})
    return {"message": "Project and related data deleted"}

# ==================== DEVICE ENDPOINTS ====================
@api_router.post("/projects/{project_id}/devices", response_model=Device)
async def create_device(project_id: str, device: DeviceBase, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    dev = Device(**device.model_dump(), project_id=project_id)
    doc = dev.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('last_poll'):
        doc['last_poll'] = doc['last_poll'].isoformat()
    await db.devices.insert_one(doc)
    await log_audit(user, "device_created", {"device_id": dev.id, "name": dev.name})
    return dev

@api_router.get("/projects/{project_id}/devices", response_model=List[Device])
async def list_devices(project_id: str, user: User = Depends(get_current_user)):
    devices = await db.devices.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    for d in devices:
        if isinstance(d.get('created_at'), str):
            d['created_at'] = datetime.fromisoformat(d['created_at'])
        if isinstance(d.get('last_poll'), str):
            d['last_poll'] = datetime.fromisoformat(d['last_poll'])
    return devices

@api_router.get("/devices/{device_id}", response_model=Device)
async def get_device(device_id: str, user: User = Depends(get_current_user)):
    device = await db.devices.find_one({"id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if isinstance(device.get('created_at'), str):
        device['created_at'] = datetime.fromisoformat(device['created_at'])
    if isinstance(device.get('last_poll'), str):
        device['last_poll'] = datetime.fromisoformat(device['last_poll'])
    return device

@api_router.put("/devices/{device_id}", response_model=Device)
async def update_device(device_id: str, update: DeviceBase, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    result = await db.devices.update_one({"id": device_id}, {"$set": update.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Device not found")
    await log_audit(user, "device_updated", {"device_id": device_id})
    return await get_device(device_id, user)

@api_router.delete("/devices/{device_id}")
async def delete_device(device_id: str, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    result = await db.devices.delete_one({"id": device_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Device not found")
    await db.tags.delete_many({"device_id": device_id})
    await log_audit(user, "device_deleted", {"device_id": device_id})
    return {"message": "Device and related tags deleted"}

# ==================== TAG ENDPOINTS ====================
@api_router.post("/projects/{project_id}/tags", response_model=Tag)
async def create_tag(project_id: str, tag: TagBase, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    t = Tag(**tag.model_dump(), project_id=project_id)
    doc = t.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('last_update'):
        doc['last_update'] = doc['last_update'].isoformat()
    await db.tags.insert_one(doc)
    return t

@api_router.get("/projects/{project_id}/tags")
async def list_tags(
    project_id: str,
    device_id: Optional[str] = None,
    permission: Optional[TagPermission] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 100,
    skip: int = 0,
    limit: int = 1000,
    user: User = Depends(get_current_user)
):
    """Get paginated tags for a project with total count for proper pagination"""
    query = {"project_id": project_id}
    if device_id:
        query["device_id"] = device_id
    if permission:
        query["permission"] = permission.value
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    
    # Calculate pagination
    if page > 0 and page_size > 0:
        actual_skip = (page - 1) * page_size
        actual_limit = page_size
    else:
        actual_skip = skip
        actual_limit = limit
    
    # Get total count for pagination
    total_count = await db.tags.count_documents(query)
    
    tags = await db.tags.find(query, {"_id": 0}).skip(actual_skip).limit(actual_limit).to_list(actual_limit)
    for t in tags:
        if isinstance(t.get('created_at'), str):
            t['created_at'] = datetime.fromisoformat(t['created_at'])
        if isinstance(t.get('last_update'), str):
            t['last_update'] = datetime.fromisoformat(t['last_update'])
    
    # Return paginated response
    total_pages = (total_count + page_size - 1) // page_size if page_size > 0 else 1
    
    return {
        "items": tags,
        "total": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "has_next": page < total_pages,
        "has_prev": page > 1
    }

@api_router.get("/tags/{tag_id}", response_model=Tag)
async def get_tag(tag_id: str, user: User = Depends(get_current_user)):
    tag = await db.tags.find_one({"id": tag_id}, {"_id": 0})
    if not tag:
        raise HTTPException(status_code=404, detail="Tag not found")
    if isinstance(tag.get('created_at'), str):
        tag['created_at'] = datetime.fromisoformat(tag['created_at'])
    if isinstance(tag.get('last_update'), str):
        tag['last_update'] = datetime.fromisoformat(tag['last_update'])
    return tag

@api_router.put("/tags/{tag_id}", response_model=Tag)
async def update_tag(tag_id: str, update: TagBase, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    result = await db.tags.update_one({"id": tag_id}, {"$set": update.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tag not found")
    return await get_tag(tag_id, user)

@api_router.delete("/tags/{tag_id}")
async def delete_tag(tag_id: str, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    result = await db.tags.delete_one({"id": tag_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tag not found")
    return {"message": "Tag deleted"}

# ==================== EXCEL IMPORT ====================
REQUIRED_COLUMNS = ['TagName', 'DeviceName', 'ObjectType', 'Address', 'DataType', 'R_W']
OBJECT_TYPE_MAP = {
    'COIL': ObjectType.COIL, '0X': ObjectType.COIL,
    'DI': ObjectType.DISCRETE_INPUT, '1X': ObjectType.DISCRETE_INPUT,
    'IR': ObjectType.INPUT_REGISTER, '3X': ObjectType.INPUT_REGISTER,
    'HR': ObjectType.HOLDING_REGISTER, '4X': ObjectType.HOLDING_REGISTER
}
DATA_TYPE_MAP = {
    'BOOL': DataType.BOOL,
    'INT16': DataType.INT16, 'UINT16': DataType.UINT16,
    'INT32': DataType.INT32, 'UINT32': DataType.UINT32,
    'FLOAT32': DataType.FLOAT32, 'FLOAT64': DataType.FLOAT64,
    'STRING': DataType.STRING
}
PERMISSION_MAP = {'R': TagPermission.READ, 'W': TagPermission.WRITE, 'RW': TagPermission.READ_WRITE}
ENDIAN_MAP = {'ABCD': EndianType.ABCD, 'CDAB': EndianType.CDAB, 'BADC': EndianType.BADC, 'DCBA': EndianType.DCBA}

@api_router.post("/projects/{project_id}/import", response_model=ImportReport)
async def import_excel(
    project_id: str,
    file: UploadFile = File(...),
    mode: str = Query("replace", pattern="^(replace|merge)$"),
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    report = ImportReport()
    
    try:
        content = await file.read()
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        
        # Validate required columns
        missing = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing:
            report.errors.append({"row": 0, "message": f"Missing required columns: {missing}"})
            report.error_count = 1
            return report
        
        # Get column indices
        col_idx = {h: i for i, h in enumerate(headers) if h}
        
        # Get device map
        devices = await db.devices.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
        device_map = {d['name']: d['id'] for d in devices}
        
        if mode == "replace":
            await db.tags.delete_many({"project_id": project_id})
        
        existing_tags = set()
        if mode == "merge":
            tags = await db.tags.find({"project_id": project_id}, {"name": 1, "_id": 0}).to_list(50000)
            existing_tags = {t['name'] for t in tags}
        
        # Track devices to auto-create
        devices_to_create = set()
        
        # First pass: identify missing devices
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[col_idx['DeviceName']] and row[col_idx['DeviceName']] not in device_map:
                devices_to_create.add(row[col_idx['DeviceName']])
        
        # Auto-create missing devices
        for device_name in devices_to_create:
            new_device = {
                "id": str(uuid.uuid4()),
                "name": device_name,
                "project_id": project_id,
                "protocol": "tcp",
                "ip_address": "127.0.0.1",
                "port": 502,
                "unit_id": 1,
                "timeout_ms": 3000,
                "retries": 3,
                "max_block_size": 120,
                "default_endian": "ABCD",
                "is_enabled": True,
                "status": "offline",
                "error_count": 0,
                "success_count": 0,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.devices.insert_one(new_device)
            device_map[device_name] = new_device["id"]
            report.warnings.append({"row": 0, "message": f"Auto-created device '{device_name}'"})
            report.warning_count += 1
        
        tags_to_insert = []
        row_num = 1
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            report.total_rows += 1
            
            try:
                tag_name = row[col_idx['TagName']]
                device_name = row[col_idx['DeviceName']]
                
                if not tag_name or not device_name:
                    report.warnings.append({"row": row_num, "message": "Empty TagName or DeviceName"})
                    report.warning_count += 1
                    continue
                
                if tag_name in existing_tags:
                    if mode == "merge":
                        report.warnings.append({"row": row_num, "message": f"Tag '{tag_name}' already exists, skipping"})
                        report.warning_count += 1
                        continue
                
                # Device should exist now (auto-created if missing)
                if device_name not in device_map:
                    report.errors.append({"row": row_num, "message": f"Device '{device_name}' could not be created"})
                    report.error_count += 1
                    continue
                
                obj_type_str = str(row[col_idx['ObjectType']]).upper()
                obj_type = OBJECT_TYPE_MAP.get(obj_type_str)
                if not obj_type:
                    report.errors.append({"row": row_num, "message": f"Invalid ObjectType: {obj_type_str}"})
                    report.error_count += 1
                    continue
                
                data_type_str = str(row[col_idx['DataType']]).upper()
                data_type = DATA_TYPE_MAP.get(data_type_str)
                if not data_type:
                    report.errors.append({"row": row_num, "message": f"Invalid DataType: {data_type_str}"})
                    report.error_count += 1
                    continue
                
                permission_str = str(row[col_idx['R_W']]).upper()
                permission = PERMISSION_MAP.get(permission_str, TagPermission.READ)
                
                tag_doc = {
                    "id": str(uuid.uuid4()),
                    "name": tag_name,
                    "project_id": project_id,
                    "device_id": device_map[device_name],
                    "object_type": obj_type.value,
                    "address": int(row[col_idx['Address']]),
                    "data_type": data_type.value,
                    "permission": permission.value,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "quality": TagQuality.UNCERTAIN.value
                }
                
                # Optional fields
                if 'Bit' in col_idx and row[col_idx['Bit']] is not None:
                    tag_doc['bit'] = int(row[col_idx['Bit']])
                if 'Scale' in col_idx and row[col_idx['Scale']] is not None:
                    tag_doc['scale'] = float(row[col_idx['Scale']])
                if 'Offset' in col_idx and row[col_idx['Offset']] is not None:
                    tag_doc['offset'] = float(row[col_idx['Offset']])
                if 'Unit' in col_idx and row[col_idx['Unit']]:
                    tag_doc['unit'] = str(row[col_idx['Unit']])
                if 'Poll_ms' in col_idx and row[col_idx['Poll_ms']] is not None:
                    tag_doc['poll_ms'] = int(row[col_idx['Poll_ms']])
                if 'Min' in col_idx and row[col_idx['Min']] is not None:
                    tag_doc['min_value'] = float(row[col_idx['Min']])
                if 'Max' in col_idx and row[col_idx['Max']] is not None:
                    tag_doc['max_value'] = float(row[col_idx['Max']])
                if 'Endian' in col_idx and row[col_idx['Endian']]:
                    endian = ENDIAN_MAP.get(str(row[col_idx['Endian']]).upper())
                    if endian:
                        tag_doc['endian'] = endian.value
                if 'Description' in col_idx and row[col_idx['Description']]:
                    tag_doc['description'] = str(row[col_idx['Description']])
                if 'AlarmEnable' in col_idx:
                    tag_doc['alarm_enable'] = str(row[col_idx['AlarmEnable']]).upper() == 'Y'
                if 'WriteConfirm' in col_idx:
                    tag_doc['write_confirm'] = str(row[col_idx['WriteConfirm']]).upper() == 'Y'
                if 'SpareFlag' in col_idx:
                    tag_doc['is_spare'] = str(row[col_idx['SpareFlag']]).upper() == 'Y'
                
                tags_to_insert.append(tag_doc)
                existing_tags.add(tag_name)
                report.success_count += 1
                
            except Exception as e:
                report.errors.append({"row": row_num, "message": str(e)})
                report.error_count += 1
        
        if tags_to_insert:
            await db.tags.insert_many(tags_to_insert)
        
        await log_audit(user, "excel_imported", {
            "project_id": project_id,
            "mode": mode,
            "total_rows": report.total_rows,
            "success": report.success_count,
            "errors": report.error_count
        })
        
    except Exception as e:
        report.errors.append({"row": 0, "message": f"Failed to parse Excel file: {str(e)}"})
        report.error_count += 1
    
    return report

@api_router.get("/projects/{project_id}/template")
async def download_template(project_id: str, user: User = Depends(get_current_user)):
    """Generate and return an Excel template"""
    from fastapi.responses import StreamingResponse
    
    output = BytesIO()
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Tag Mapping"
    
    headers = [
        'TagName', 'DeviceName', 'ObjectType', 'Address', 'Bit', 'DataType',
        'R_W', 'Scale', 'Offset', 'Unit', 'Poll_ms', 'Min', 'Max',
        'Endian', 'Description', 'AlarmEnable', 'Deadband', 'WriteConfirm', 'SpareFlag'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add sample data
    sample = [
        'VENDOR_A_PT_101', 'VendorA_PLC', 'HR', 40001, '', 'FLOAT32',
        'R', 0.1, 0, 'bar', 1000, 0, 100, 'ABCD', 'Suction Pressure', 'Y', 0.5, 'N', 'N'
    ]
    for col, value in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=value)
    
    workbook.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tag_mapping_template.xlsx"}
    )

# ==================== TRAFFIC LOGS ====================
@api_router.get("/projects/{project_id}/traffic", response_model=List[TrafficLog])
async def get_traffic_logs(
    project_id: str,
    device_id: Optional[str] = None,
    status: Optional[TrafficStatus] = None,
    limit: int = 100,
    user: User = Depends(get_current_user)
):
    query = {"project_id": project_id}
    if device_id:
        query["device_id"] = device_id
    if status:
        query["status"] = status.value
    
    logs = await db.traffic_logs.find(query, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    for log in logs:
        if isinstance(log.get('timestamp'), str):
            log['timestamp'] = datetime.fromisoformat(log['timestamp'])
    return logs

@api_router.delete("/projects/{project_id}/traffic")
async def clear_traffic_logs(project_id: str, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    await db.traffic_logs.delete_many({"project_id": project_id})
    return {"message": "Traffic logs cleared"}

@api_router.get("/projects/{project_id}/traffic/export")
async def export_traffic_logs(project_id: str, user: User = Depends(get_current_user)):
    from fastapi.responses import StreamingResponse
    import csv
    
    logs = await db.traffic_logs.find({"project_id": project_id}, {"_id": 0}).sort("timestamp", -1).to_list(10000)
    
    output = BytesIO()
    # Write CSV with proper encoding
    
    fieldnames = ['timestamp', 'device_name', 'protocol', 'function_code', 'request_summary', 
                  'response_summary', 'round_trip_ms', 'status', 'error_details']
    
    import io
    text_output = io.StringIO()
    writer = csv.DictWriter(text_output, fieldnames=fieldnames)
    writer.writeheader()
    
    for log in logs:
        writer.writerow({
            'timestamp': log.get('timestamp', ''),
            'device_name': log.get('device_name', ''),
            'protocol': log.get('protocol', ''),
            'function_code': log.get('function_code', ''),
            'request_summary': log.get('request_summary', ''),
            'response_summary': log.get('response_summary', ''),
            'round_trip_ms': log.get('round_trip_ms', ''),
            'status': log.get('status', ''),
            'error_details': log.get('error_details', '')
        })
    
    output = BytesIO(text_output.getvalue().encode('utf-8'))
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=traffic_log_{project_id}.csv"}
    )

# ==================== AUDIT LOGS ====================
@api_router.get("/audit-logs", response_model=List[AuditLog])
async def get_audit_logs(
    limit: int = 100,
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    logs = await db.audit_logs.find({}, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    for log in logs:
        if isinstance(log.get('timestamp'), str):
            log['timestamp'] = datetime.fromisoformat(log['timestamp'])
    return logs

# ==================== MODBUS SIMULATOR ====================
simulator_sessions = {}
simulator_data = {}

class ModbusSimulator:
    def __init__(self, session_id: str, config: SimulatorConfig):
        self.session_id = session_id
        self.config = config
        self.coils = [False] * config.num_coils
        self.discrete_inputs = [False] * config.num_discrete_inputs
        self.input_registers = [0] * config.num_input_registers
        self.holding_registers = [0] * config.num_holding_registers
        self.is_running = False
        
    def read_coils(self, address: int, count: int) -> List[bool]:
        return self.coils[address:address + count]
    
    def read_discrete_inputs(self, address: int, count: int) -> List[bool]:
        return self.discrete_inputs[address:address + count]
    
    def read_input_registers(self, address: int, count: int) -> List[int]:
        return self.input_registers[address:address + count]
    
    def read_holding_registers(self, address: int, count: int) -> List[int]:
        return self.holding_registers[address:address + count]
    
    def write_coil(self, address: int, value: bool):
        if 0 <= address < len(self.coils):
            self.coils[address] = value
    
    def write_register(self, address: int, value: int):
        if 0 <= address < len(self.holding_registers):
            self.holding_registers[address] = value & 0xFFFF
    
    def write_multiple_coils(self, address: int, values: List[bool]):
        for i, val in enumerate(values):
            if address + i < len(self.coils):
                self.coils[address + i] = val
    
    def write_multiple_registers(self, address: int, values: List[int]):
        for i, val in enumerate(values):
            if address + i < len(self.holding_registers):
                self.holding_registers[address + i] = val & 0xFFFF

@api_router.post("/simulator/start", response_model=SimulatorSession)
async def start_simulator(config: SimulatorConfig, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    session = SimulatorSession(config=config, is_running=True)
    simulator = ModbusSimulator(session.id, config)
    simulator.is_running = True
    simulator_sessions[session.id] = simulator
    
    await log_audit(user, "simulator_started", {"session_id": session.id, "protocol": config.protocol.value})
    return session

@api_router.post("/simulator/{session_id}/stop")
async def stop_simulator(session_id: str, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    if session_id not in simulator_sessions:
        raise HTTPException(status_code=404, detail="Simulator session not found")
    
    simulator_sessions[session_id].is_running = False
    del simulator_sessions[session_id]
    
    await log_audit(user, "simulator_stopped", {"session_id": session_id})
    return {"message": "Simulator stopped"}

@api_router.get("/simulator/sessions")
async def list_simulator_sessions(user: User = Depends(get_current_user)):
    return [
        {
            "id": sid,
            "is_running": sim.is_running,
            "protocol": sim.config.protocol.value,
            "port": sim.config.port
        }
        for sid, sim in simulator_sessions.items()
    ]

@api_router.get("/simulator/{session_id}/data")
async def get_simulator_data(session_id: str, user: User = Depends(get_current_user)):
    if session_id not in simulator_sessions:
        raise HTTPException(status_code=404, detail="Simulator session not found")
    
    sim = simulator_sessions[session_id]
    return {
        "coils": sim.coils[:20],  # Return first 20 for preview
        "discrete_inputs": sim.discrete_inputs[:20],
        "input_registers": sim.input_registers[:20],
        "holding_registers": sim.holding_registers[:20]
    }

@api_router.post("/simulator/{session_id}/write")
async def write_simulator_data(
    session_id: str,
    object_type: ObjectType,
    address: int,
    value: Any,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER, UserRole.OPERATOR]))
):
    if session_id not in simulator_sessions:
        raise HTTPException(status_code=404, detail="Simulator session not found")
    
    sim = simulator_sessions[session_id]
    
    if object_type == ObjectType.COIL:
        sim.write_coil(address, bool(value))
    elif object_type == ObjectType.HOLDING_REGISTER:
        sim.write_register(address, int(value))
    else:
        raise HTTPException(status_code=400, detail="Can only write to coils or holding registers")
    
    return {"message": "Write successful"}

# ==================== MODBUS SERVER (SLAVE MODE) ====================
# Real network-capable Modbus server using pymodbus
from pymodbus.datastore import ModbusSequentialDataBlock, ModbusDeviceContext, ModbusServerContext
from pymodbus.server import StartAsyncTcpServer, StartAsyncUdpServer

# Global storage for server instances
modbus_servers = {}

class ModbusServerConfig(BaseModel):
    protocol: Literal["tcp", "udp"] = "tcp"
    host: str = "0.0.0.0"
    port: int = 5020
    unit_id: int = 1
    num_coils: int = 1000
    num_discrete_inputs: int = 1000
    num_input_registers: int = 1000
    num_holding_registers: int = 1000

class ModbusServerInfo(BaseModel):
    id: str
    protocol: str
    host: str
    port: int
    unit_id: int
    is_running: bool
    created_at: str

class ModbusServerInstance:
    def __init__(self, server_id: str, config: ModbusServerConfig):
        self.id = server_id
        self.config = config
        self.is_running = False
        self.task = None
        self.created_at = datetime.now(timezone.utc)
        
        # Create data blocks
        self.coils_block = ModbusSequentialDataBlock(0, [False] * config.num_coils)
        self.discrete_block = ModbusSequentialDataBlock(0, [False] * config.num_discrete_inputs)
        self.input_block = ModbusSequentialDataBlock(0, [0] * config.num_input_registers)
        self.holding_block = ModbusSequentialDataBlock(0, [0] * config.num_holding_registers)
        
        # Create slave context
        self.slave_context = ModbusDeviceContext(
            di=self.discrete_block,
            co=self.coils_block,
            hr=self.holding_block,
            ir=self.input_block
        )
        
        # Create server context
        self.server_context = ModbusServerContext(
            devices={config.unit_id: self.slave_context},
            single=False
        )
    
    async def start(self):
        if self.is_running:
            return
        
        self.is_running = True
        self.task = asyncio.create_task(self._run_server())
        logger.info(f"Modbus server {self.id} started on {self.config.host}:{self.config.port} ({self.config.protocol.upper()})")
    
    async def _run_server(self):
        try:
            address = (self.config.host, self.config.port)
            
            if self.config.protocol == "tcp":
                await StartAsyncTcpServer(
                    context=self.server_context,
                    address=address
                )
            else:  # UDP
                await StartAsyncUdpServer(
                    context=self.server_context,
                    address=address
                )
        except asyncio.CancelledError:
            logger.info(f"Modbus server {self.id} cancelled")
        except Exception as e:
            logger.error(f"Modbus server {self.id} error: {e}")
            self.is_running = False
    
    async def stop(self):
        if not self.is_running:
            return
        
        self.is_running = False
        if self.task:
            self.task.cancel()
            try:
                await self.task
            except asyncio.CancelledError:
                pass
        logger.info(f"Modbus server {self.id} stopped")
    
    def read_registers(self, register_type: str, address: int, count: int) -> List[int]:
        """Read register values from the datastore"""
        try:
            if register_type == "coil":
                return self.coils_block.getValues(address, count)
            elif register_type == "discrete_input":
                return self.discrete_block.getValues(address, count)
            elif register_type == "input_register":
                return self.input_block.getValues(address, count)
            elif register_type == "holding_register":
                return self.holding_block.getValues(address, count)
            return []
        except Exception as e:
            logger.error(f"Read error: {e}")
            return []
    
    def write_registers(self, register_type: str, address: int, values: List[Any]) -> bool:
        """Write register values to the datastore"""
        try:
            if register_type == "coil":
                self.coils_block.setValues(address, values)
            elif register_type == "holding_register":
                self.holding_block.setValues(address, values)
            elif register_type == "input_register":
                self.input_block.setValues(address, values)
            elif register_type == "discrete_input":
                self.discrete_block.setValues(address, values)
            return True
        except Exception as e:
            logger.error(f"Write error: {e}")
            return False

@api_router.post("/modbus-server/start", response_model=ModbusServerInfo)
async def start_modbus_server(
    config: ModbusServerConfig,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Start a new Modbus server (slave) that can accept connections from other devices/PCs"""
    
    # Check if port is already in use
    for srv in modbus_servers.values():
        if srv.is_running and srv.config.port == config.port:
            raise HTTPException(
                status_code=400, 
                detail=f"Port {config.port} is already in use by another Modbus server"
            )
    
    server_id = str(uuid.uuid4())
    server = ModbusServerInstance(server_id, config)
    modbus_servers[server_id] = server
    
    await server.start()
    await log_audit(user, "modbus_server_started", {
        "server_id": server_id,
        "protocol": config.protocol,
        "port": config.port
    })
    
    return ModbusServerInfo(
        id=server_id,
        protocol=config.protocol,
        host=config.host,
        port=config.port,
        unit_id=config.unit_id,
        is_running=True,
        created_at=server.created_at.isoformat()
    )

@api_router.post("/modbus-server/{server_id}/stop")
async def stop_modbus_server(
    server_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Stop a running Modbus server"""
    if server_id not in modbus_servers:
        raise HTTPException(status_code=404, detail="Modbus server not found")
    
    server = modbus_servers[server_id]
    await server.stop()
    del modbus_servers[server_id]
    
    await log_audit(user, "modbus_server_stopped", {"server_id": server_id})
    return {"message": "Modbus server stopped"}

@api_router.get("/modbus-server/list")
async def list_modbus_servers(user: User = Depends(get_current_user)):
    """List all active Modbus servers"""
    return [
        ModbusServerInfo(
            id=srv.id,
            protocol=srv.config.protocol,
            host=srv.config.host,
            port=srv.config.port,
            unit_id=srv.config.unit_id,
            is_running=srv.is_running,
            created_at=srv.created_at.isoformat()
        )
        for srv in modbus_servers.values()
    ]

@api_router.get("/modbus-server/{server_id}/data")
async def get_modbus_server_data(
    server_id: str,
    user: User = Depends(get_current_user)
):
    """Get current register values from a Modbus server"""
    if server_id not in modbus_servers:
        raise HTTPException(status_code=404, detail="Modbus server not found")
    
    server = modbus_servers[server_id]
    return {
        "coils": server.read_registers("coil", 0, 20),
        "discrete_inputs": server.read_registers("discrete_input", 0, 20),
        "input_registers": server.read_registers("input_register", 0, 20),
        "holding_registers": server.read_registers("holding_register", 0, 20)
    }

class ServerWriteRequest(BaseModel):
    register_type: Literal["coil", "holding_register", "input_register", "discrete_input"]
    address: int
    values: List[Any]

@api_router.post("/modbus-server/{server_id}/write")
async def write_modbus_server_data(
    server_id: str,
    request: ServerWriteRequest,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER, UserRole.OPERATOR]))
):
    """Write values to a Modbus server's registers"""
    if server_id not in modbus_servers:
        raise HTTPException(status_code=404, detail="Modbus server not found")
    
    server = modbus_servers[server_id]
    success = server.write_registers(request.register_type, request.address, request.values)
    
    if not success:
        raise HTTPException(status_code=400, detail="Failed to write values")
    
    await log_audit(user, "modbus_server_write", {
        "server_id": server_id,
        "register_type": request.register_type,
        "address": request.address,
        "count": len(request.values)
    })
    
    return {"message": "Write successful", "address": request.address, "count": len(request.values)}

# ==================== DATA SIMULATION ENGINE ====================
import math

class SimulationConfig(BaseModel):
    enabled: bool = True
    interval_ms: int = 1000
    pattern: Literal["sine", "ramp", "random", "square"] = "sine"
    amplitude: float = 50.0
    offset: float = 50.0
    period_seconds: float = 60.0  # One full cycle in seconds

simulation_tasks = {}

class DataSimulationEngine:
    """Generates realistic sine wave data for Modbus server registers"""
    
    def __init__(self, server_id: str, config: SimulationConfig):
        self.server_id = server_id
        self.config = config
        self.is_running = False
        self.task = None
        self.start_time = datetime.now(timezone.utc)
    
    async def start(self):
        if self.is_running:
            return
        self.is_running = True
        self.start_time = datetime.now(timezone.utc)
        self.task = asyncio.create_task(self._simulation_loop())
        logger.info(f"Data simulation started for server {self.server_id}")
    
    async def stop(self):
        self.is_running = False
        if self.task:
            self.task.cancel()
            try:
                await self.task
            except asyncio.CancelledError:
                pass
        logger.info(f"Data simulation stopped for server {self.server_id}")
    
    def _generate_value(self, register_index: int) -> int:
        """Generate a simulated value based on pattern"""
        elapsed = (datetime.now(timezone.utc) - self.start_time).total_seconds()
        # Add phase offset per register for variety
        phase_offset = register_index * (math.pi / 10)
        
        if self.config.pattern == "sine":
            # Sine wave: value = offset + amplitude * sin(2*pi*t/period + phase)
            angle = (2 * math.pi * elapsed / self.config.period_seconds) + phase_offset
            value = self.config.offset + self.config.amplitude * math.sin(angle)
        elif self.config.pattern == "ramp":
            # Sawtooth ramp
            cycle_position = (elapsed % self.config.period_seconds) / self.config.period_seconds
            value = self.config.offset - self.config.amplitude + (2 * self.config.amplitude * cycle_position)
        elif self.config.pattern == "square":
            # Square wave
            cycle_position = (elapsed % self.config.period_seconds) / self.config.period_seconds
            value = self.config.offset + self.config.amplitude if cycle_position < 0.5 else self.config.offset - self.config.amplitude
        else:  # random
            import random
            value = self.config.offset + (random.random() - 0.5) * 2 * self.config.amplitude
        
        # Clamp to valid 16-bit unsigned range
        return max(0, min(65535, int(value)))
    
    async def _simulation_loop(self):
        """Main simulation loop that updates server registers"""
        try:
            while self.is_running:
                if self.server_id in modbus_servers:
                    server = modbus_servers[self.server_id]
                    
                    # Update holding registers with simulated values
                    for i in range(20):  # Update first 20 registers
                        value = self._generate_value(i)
                        server.write_registers("holding_register", i, [value])
                    
                    # Update input registers similarly
                    for i in range(20):
                        value = self._generate_value(i + 20)  # Different phase
                        server.write_registers("input_register", i, [value])
                    
                    # Update some coils based on threshold
                    for i in range(10):
                        value = self._generate_value(i)
                        server.write_registers("coil", i, [value > (self.config.offset)])
                
                await asyncio.sleep(self.config.interval_ms / 1000.0)
                
        except asyncio.CancelledError:
            pass
        except Exception as e:
            logger.error(f"Simulation error: {e}")

@api_router.post("/modbus-server/{server_id}/simulation/start")
async def start_simulation(
    server_id: str,
    config: SimulationConfig,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Start data simulation for a Modbus server (generates sine wave patterns)"""
    if server_id not in modbus_servers:
        raise HTTPException(status_code=404, detail="Modbus server not found")
    
    # Stop existing simulation if any
    if server_id in simulation_tasks:
        await simulation_tasks[server_id].stop()
    
    sim = DataSimulationEngine(server_id, config)
    simulation_tasks[server_id] = sim
    await sim.start()
    
    await log_audit(user, "simulation_started", {
        "server_id": server_id,
        "pattern": config.pattern,
        "interval_ms": config.interval_ms
    })
    
    return {
        "message": "Simulation started",
        "server_id": server_id,
        "pattern": config.pattern,
        "interval_ms": config.interval_ms
    }

@api_router.post("/modbus-server/{server_id}/simulation/stop")
async def stop_simulation(
    server_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Stop data simulation for a Modbus server"""
    if server_id not in simulation_tasks:
        raise HTTPException(status_code=404, detail="No simulation running for this server")
    
    await simulation_tasks[server_id].stop()
    del simulation_tasks[server_id]
    
    await log_audit(user, "simulation_stopped", {"server_id": server_id})
    return {"message": "Simulation stopped"}

@api_router.get("/modbus-server/{server_id}/simulation/status")
async def get_simulation_status(
    server_id: str,
    user: User = Depends(get_current_user)
):
    """Get simulation status for a Modbus server"""
    if server_id in simulation_tasks:
        sim = simulation_tasks[server_id]
        return {
            "is_running": sim.is_running,
            "pattern": sim.config.pattern,
            "interval_ms": sim.config.interval_ms,
            "amplitude": sim.config.amplitude,
            "offset": sim.config.offset,
            "period_seconds": sim.config.period_seconds
        }
    return {"is_running": False}

# ==================== MODBUS CLIENT (MASTER MODE) ====================
from pymodbus.client import AsyncModbusTcpClient, AsyncModbusUdpClient

modbus_clients = {}

class ModbusClientConfig(BaseModel):
    protocol: Literal["tcp", "udp"] = "tcp"
    remote_host: str  # IP address of remote Modbus server
    remote_port: int = 5020
    unit_id: int = 1
    poll_interval_ms: int = 1000
    read_coils: bool = True
    read_discrete_inputs: bool = True
    read_input_registers: bool = True
    read_holding_registers: bool = True
    start_address: int = 0
    register_count: int = 20

class ModbusClientInfo(BaseModel):
    id: str
    protocol: str
    remote_host: str
    remote_port: int
    unit_id: int
    is_connected: bool
    poll_interval_ms: int
    created_at: str
    last_poll: Optional[str] = None
    error_message: Optional[str] = None

class ModbusClientInstance:
    """Modbus client that connects to remote servers and reads data"""
    
    def __init__(self, client_id: str, config: ModbusClientConfig):
        self.id = client_id
        self.config = config
        self.is_connected = False
        self.is_running = False
        self.task = None
        self.client = None
        self.created_at = datetime.now(timezone.utc)
        self.last_poll = None
        self.error_message = None
        
        # Store received data
        self.received_data = {
            "coils": [],
            "discrete_inputs": [],
            "input_registers": [],
            "holding_registers": []
        }
    
    async def connect(self):
        """Establish connection to remote Modbus server"""
        try:
            if self.config.protocol == "tcp":
                self.client = AsyncModbusTcpClient(
                    host=self.config.remote_host,
                    port=self.config.remote_port
                )
            else:
                self.client = AsyncModbusUdpClient(
                    host=self.config.remote_host,
                    port=self.config.remote_port
                )
            
            connected = await self.client.connect()
            self.is_connected = connected
            if connected:
                logger.info(f"Modbus client {self.id} connected to {self.config.remote_host}:{self.config.remote_port}")
                self.error_message = None
            else:
                self.error_message = "Failed to connect"
                logger.error(f"Modbus client {self.id} failed to connect")
            return connected
        except Exception as e:
            self.error_message = str(e)
            logger.error(f"Modbus client {self.id} connection error: {e}")
            return False
    
    async def disconnect(self):
        """Disconnect from remote server"""
        if self.client:
            self.client.close()
        self.is_connected = False
        logger.info(f"Modbus client {self.id} disconnected")
    
    async def start_polling(self):
        """Start the polling loop"""
        if self.is_running:
            return
        
        self.is_running = True
        self.task = asyncio.create_task(self._poll_loop())
    
    async def stop_polling(self):
        """Stop the polling loop"""
        self.is_running = False
        if self.task:
            self.task.cancel()
            try:
                await self.task
            except asyncio.CancelledError:
                pass
        await self.disconnect()
    
    async def _poll_loop(self):
        """Main polling loop - reads data from remote server"""
        try:
            while self.is_running:
                if not self.is_connected:
                    await self.connect()
                    if not self.is_connected:
                        await asyncio.sleep(5)  # Retry connection after 5 seconds
                        continue
                
                try:
                    # Read coils
                    if self.config.read_coils:
                        result = await self.client.read_coils(
                            address=self.config.start_address,
                            count=self.config.register_count,
                            device_id=self.config.unit_id
                        )
                        if not result.isError():
                            self.received_data["coils"] = list(result.bits[:self.config.register_count])
                    
                    # Read discrete inputs
                    if self.config.read_discrete_inputs:
                        result = await self.client.read_discrete_inputs(
                            address=self.config.start_address,
                            count=self.config.register_count,
                            device_id=self.config.unit_id
                        )
                        if not result.isError():
                            self.received_data["discrete_inputs"] = list(result.bits[:self.config.register_count])
                    
                    # Read input registers
                    if self.config.read_input_registers:
                        result = await self.client.read_input_registers(
                            address=self.config.start_address,
                            count=self.config.register_count,
                            device_id=self.config.unit_id
                        )
                        if not result.isError():
                            self.received_data["input_registers"] = list(result.registers)
                    
                    # Read holding registers
                    if self.config.read_holding_registers:
                        result = await self.client.read_holding_registers(
                            address=self.config.start_address,
                            count=self.config.register_count,
                            device_id=self.config.unit_id
                        )
                        if not result.isError():
                            self.received_data["holding_registers"] = list(result.registers)
                    
                    self.last_poll = datetime.now(timezone.utc)
                    self.error_message = None
                    
                except Exception as e:
                    self.error_message = str(e)
                    logger.error(f"Modbus client {self.id} poll error: {e}")
                    self.is_connected = False
                
                await asyncio.sleep(self.config.poll_interval_ms / 1000.0)
                
        except asyncio.CancelledError:
            pass
        except Exception as e:
            logger.error(f"Modbus client {self.id} loop error: {e}")

@api_router.post("/modbus-client/connect")
async def connect_modbus_client(
    config: ModbusClientConfig,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Connect to a remote Modbus server as a client (master)"""
    client_id = str(uuid.uuid4())
    client = ModbusClientInstance(client_id, config)
    modbus_clients[client_id] = client
    
    # Connect and start polling
    connected = await client.connect()
    if connected:
        await client.start_polling()
    
    await log_audit(user, "modbus_client_connected", {
        "client_id": client_id,
        "remote_host": config.remote_host,
        "remote_port": config.remote_port
    })
    
    return ModbusClientInfo(
        id=client_id,
        protocol=config.protocol,
        remote_host=config.remote_host,
        remote_port=config.remote_port,
        unit_id=config.unit_id,
        is_connected=client.is_connected,
        poll_interval_ms=config.poll_interval_ms,
        created_at=client.created_at.isoformat(),
        error_message=client.error_message
    )

@api_router.post("/modbus-client/{client_id}/disconnect")
async def disconnect_modbus_client(
    client_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Disconnect a Modbus client"""
    if client_id not in modbus_clients:
        raise HTTPException(status_code=404, detail="Modbus client not found")
    
    client = modbus_clients[client_id]
    await client.stop_polling()
    del modbus_clients[client_id]
    
    await log_audit(user, "modbus_client_disconnected", {"client_id": client_id})
    return {"message": "Client disconnected"}

@api_router.get("/modbus-client/list")
async def list_modbus_clients(user: User = Depends(get_current_user)):
    """List all active Modbus clients"""
    return [
        ModbusClientInfo(
            id=client.id,
            protocol=client.config.protocol,
            remote_host=client.config.remote_host,
            remote_port=client.config.remote_port,
            unit_id=client.config.unit_id,
            is_connected=client.is_connected,
            poll_interval_ms=client.config.poll_interval_ms,
            created_at=client.created_at.isoformat(),
            last_poll=client.last_poll.isoformat() if client.last_poll else None,
            error_message=client.error_message
        )
        for client in modbus_clients.values()
    ]

@api_router.get("/modbus-client/{client_id}/data")
async def get_modbus_client_data(
    client_id: str,
    user: User = Depends(get_current_user)
):
    """Get data received from remote Modbus server"""
    if client_id not in modbus_clients:
        raise HTTPException(status_code=404, detail="Modbus client not found")
    
    client = modbus_clients[client_id]
    return {
        "is_connected": client.is_connected,
        "last_poll": client.last_poll.isoformat() if client.last_poll else None,
        "error_message": client.error_message,
        "data": client.received_data
    }

@api_router.post("/modbus-client/{client_id}/write")
async def write_via_modbus_client(
    client_id: str,
    request: ServerWriteRequest,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER, UserRole.OPERATOR]))
):
    """Write values to remote Modbus server via client"""
    if client_id not in modbus_clients:
        raise HTTPException(status_code=404, detail="Modbus client not found")
    
    client = modbus_clients[client_id]
    if not client.is_connected or not client.client:
        raise HTTPException(status_code=400, detail="Client not connected")
    
    try:
        if request.register_type == "coil":
            if len(request.values) == 1:
                result = await client.client.write_coil(
                    request.address,
                    request.values[0],
                    device_id=client.config.unit_id
                )
            else:
                result = await client.client.write_coils(
                    request.address,
                    request.values,
                    device_id=client.config.unit_id
                )
        elif request.register_type == "holding_register":
            if len(request.values) == 1:
                result = await client.client.write_register(
                    request.address,
                    request.values[0],
                    device_id=client.config.unit_id
                )
            else:
                result = await client.client.write_registers(
                    request.address,
                    request.values,
                    device_id=client.config.unit_id
                )
        else:
            raise HTTPException(status_code=400, detail="Can only write to coils or holding registers")
        
        if result.isError():
            raise HTTPException(status_code=400, detail=f"Write failed: {result}")
        
        await log_audit(user, "modbus_client_write", {
            "client_id": client_id,
            "register_type": request.register_type,
            "address": request.address
        })
        
        return {"message": "Write successful", "address": request.address, "count": len(request.values)}
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Write error: {str(e)}")

# ==================== OPC UA SERVER & CLIENT ====================
from asyncua import Server as OPCUAServer, Client as OPCUAClient, ua
from asyncua.crypto.security_policies import SecurityPolicyBasic256Sha256

opcua_servers = {}
opcua_clients = {}

class OPCUASecurityMode(str, Enum):
    NONE = "none"
    SIGN = "sign"
    SIGN_ENCRYPT = "sign_encrypt"

class OPCUAServerConfig(BaseModel):
    endpoint_url: str = "opc.tcp://0.0.0.0:4840/comgate/server/"
    server_name: str = "ComGate OPC UA Server"
    namespace: str = "http://comgate.industrial/opcua"
    security_mode: OPCUASecurityMode = OPCUASecurityMode.NONE
    num_variables: int = 100

class OPCUAServerInfo(BaseModel):
    id: str
    endpoint_url: str
    server_name: str
    namespace: str
    security_mode: str
    is_running: bool
    simulation_enabled: bool
    created_at: str

class OPCUAServerInstance:
    """OPC UA Server that exposes tags to OPC UA clients"""
    
    def __init__(self, server_id: str, config: OPCUAServerConfig):
        self.id = server_id
        self.config = config
        self.server = None
        self.is_running = False
        self.simulation_enabled = False
        self.simulation_task = None
        self.created_at = datetime.now(timezone.utc)
        self.variables = {}  # NodeId -> variable node
        self.namespace_idx = None
    
    async def start(self):
        if self.is_running:
            return
        
        try:
            self.server = OPCUAServer()
            self.server.set_endpoint(self.config.endpoint_url)
            self.server.set_server_name(self.config.server_name)
            
            await self.server.init()
            
            # Register namespace
            self.namespace_idx = await self.server.register_namespace(self.config.namespace)
            
            # Create objects folder structure
            objects = self.server.nodes.objects
            comgate_folder = await objects.add_folder(self.namespace_idx, "ComGate")
            
            # Create variables
            for i in range(self.config.num_variables):
                var = await comgate_folder.add_variable(
                    self.namespace_idx,
                    f"Tag_{i:04d}",
                    0.0
                )
                await var.set_writable()
                self.variables[f"Tag_{i:04d}"] = var
            
            # Create status variables
            self.status_var = await comgate_folder.add_variable(
                self.namespace_idx,
                "ServerStatus",
                "Running"
            )
            self.timestamp_var = await comgate_folder.add_variable(
                self.namespace_idx,
                "LastUpdate",
                datetime.now(timezone.utc).isoformat()
            )
            
            await self.server.start()
            self.is_running = True
            logger.info(f"OPC UA Server {self.id} started at {self.config.endpoint_url}")
            
        except Exception as e:
            logger.error(f"Failed to start OPC UA server: {e}")
            raise
    
    async def stop(self):
        if not self.is_running:
            return
        
        if self.simulation_task:
            self.simulation_task.cancel()
            try:
                await self.simulation_task
            except asyncio.CancelledError:
                pass
        
        if self.server:
            await self.server.stop()
        
        self.is_running = False
        logger.info(f"OPC UA Server {self.id} stopped")
    
    async def start_simulation(self, interval_ms: int = 1000, pattern: str = "sine"):
        if self.simulation_enabled:
            return
        
        self.simulation_enabled = True
        self.simulation_task = asyncio.create_task(self._simulation_loop(interval_ms, pattern))
    
    async def stop_simulation(self):
        self.simulation_enabled = False
        if self.simulation_task:
            self.simulation_task.cancel()
            try:
                await self.simulation_task
            except asyncio.CancelledError:
                pass
    
    async def _simulation_loop(self, interval_ms: int, pattern: str):
        start_time = datetime.now(timezone.utc)
        try:
            while self.simulation_enabled:
                elapsed = (datetime.now(timezone.utc) - start_time).total_seconds()
                
                for i, (name, var) in enumerate(self.variables.items()):
                    phase_offset = i * (math.pi / 10)
                    
                    if pattern == "sine":
                        value = 50.0 + 50.0 * math.sin((2 * math.pi * elapsed / 60.0) + phase_offset)
                    elif pattern == "ramp":
                        cycle_pos = (elapsed % 60.0) / 60.0
                        value = 100.0 * cycle_pos
                    else:
                        import random
                        value = random.uniform(0, 100)
                    
                    await var.write_value(float(value))
                
                await self.timestamp_var.write_value(datetime.now(timezone.utc).isoformat())
                await asyncio.sleep(interval_ms / 1000.0)
                
        except asyncio.CancelledError:
            pass
    
    async def read_variable(self, name: str):
        if name in self.variables:
            return await self.variables[name].read_value()
        return None
    
    async def write_variable(self, name: str, value: float):
        if name in self.variables:
            await self.variables[name].write_value(value)
            return True
        return False
    
    async def get_all_values(self) -> dict:
        result = {}
        for name, var in list(self.variables.items())[:20]:  # First 20 for display
            try:
                result[name] = await var.read_value()
            except:
                result[name] = None
        return result

@api_router.post("/opcua-server/start")
async def start_opcua_server(
    config: OPCUAServerConfig,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Start an OPC UA server"""
    server_id = str(uuid.uuid4())
    server = OPCUAServerInstance(server_id, config)
    
    try:
        await server.start()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to start server: {str(e)}")
    
    opcua_servers[server_id] = server
    
    await log_audit(user, "opcua_server_started", {
        "server_id": server_id,
        "endpoint": config.endpoint_url
    })
    
    return OPCUAServerInfo(
        id=server_id,
        endpoint_url=config.endpoint_url,
        server_name=config.server_name,
        namespace=config.namespace,
        security_mode=config.security_mode,
        is_running=True,
        simulation_enabled=False,
        created_at=server.created_at.isoformat()
    )

@api_router.post("/opcua-server/{server_id}/stop")
async def stop_opcua_server(
    server_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Stop an OPC UA server"""
    if server_id not in opcua_servers:
        raise HTTPException(status_code=404, detail="OPC UA server not found")
    
    await opcua_servers[server_id].stop()
    del opcua_servers[server_id]
    
    await log_audit(user, "opcua_server_stopped", {"server_id": server_id})
    return {"message": "OPC UA server stopped"}

@api_router.get("/opcua-server/list")
async def list_opcua_servers(user: User = Depends(get_current_user)):
    """List all active OPC UA servers"""
    return [
        OPCUAServerInfo(
            id=srv.id,
            endpoint_url=srv.config.endpoint_url,
            server_name=srv.config.server_name,
            namespace=srv.config.namespace,
            security_mode=srv.config.security_mode,
            is_running=srv.is_running,
            simulation_enabled=srv.simulation_enabled,
            created_at=srv.created_at.isoformat()
        )
        for srv in opcua_servers.values()
    ]

@api_router.get("/opcua-server/{server_id}/data")
async def get_opcua_server_data(
    server_id: str,
    user: User = Depends(get_current_user)
):
    """Get OPC UA server variable values"""
    if server_id not in opcua_servers:
        raise HTTPException(status_code=404, detail="OPC UA server not found")
    
    server = opcua_servers[server_id]
    values = await server.get_all_values()
    
    return {
        "is_running": server.is_running,
        "simulation_enabled": server.simulation_enabled,
        "variables": values
    }

@api_router.post("/opcua-server/{server_id}/simulation/start")
async def start_opcua_simulation(
    server_id: str,
    config: SimulationConfig,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Start data simulation for OPC UA server"""
    if server_id not in opcua_servers:
        raise HTTPException(status_code=404, detail="OPC UA server not found")
    
    server = opcua_servers[server_id]
    await server.start_simulation(config.interval_ms, config.pattern)
    
    return {"message": "Simulation started", "pattern": config.pattern}

@api_router.post("/opcua-server/{server_id}/simulation/stop")
async def stop_opcua_simulation(
    server_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Stop data simulation for OPC UA server"""
    if server_id not in opcua_servers:
        raise HTTPException(status_code=404, detail="OPC UA server not found")
    
    await opcua_servers[server_id].stop_simulation()
    return {"message": "Simulation stopped"}

# OPC UA Client
class OPCUAClientConfig(BaseModel):
    endpoint_url: str  # e.g., "opc.tcp://192.168.1.100:4840/server/"
    security_mode: OPCUASecurityMode = OPCUASecurityMode.NONE
    poll_interval_ms: int = 1000
    node_ids: List[str] = []  # List of node IDs to read

class OPCUAClientInfo(BaseModel):
    id: str
    endpoint_url: str
    security_mode: str
    is_connected: bool
    poll_interval_ms: int
    created_at: str
    last_poll: Optional[str] = None
    error_message: Optional[str] = None

class OPCUAClientInstance:
    """OPC UA Client that connects to remote OPC UA servers"""
    
    def __init__(self, client_id: str, config: OPCUAClientConfig):
        self.id = client_id
        self.config = config
        self.client = None
        self.is_connected = False
        self.is_running = False
        self.task = None
        self.created_at = datetime.now(timezone.utc)
        self.last_poll = None
        self.error_message = None
        self.received_data = {}
    
    async def connect(self):
        try:
            self.client = OPCUAClient(self.config.endpoint_url)
            await self.client.connect()
            self.is_connected = True
            self.error_message = None
            logger.info(f"OPC UA Client {self.id} connected to {self.config.endpoint_url}")
            return True
        except Exception as e:
            self.error_message = str(e)
            logger.error(f"OPC UA Client {self.id} connection failed: {e}")
            return False
    
    async def disconnect(self):
        if self.client:
            await self.client.disconnect()
        self.is_connected = False
    
    async def start_polling(self):
        if self.is_running:
            return
        
        self.is_running = True
        self.task = asyncio.create_task(self._poll_loop())
    
    async def stop_polling(self):
        self.is_running = False
        if self.task:
            self.task.cancel()
            try:
                await self.task
            except asyncio.CancelledError:
                pass
        await self.disconnect()
    
    async def _poll_loop(self):
        try:
            while self.is_running:
                if not self.is_connected:
                    await self.connect()
                    if not self.is_connected:
                        await asyncio.sleep(5)
                        continue
                
                try:
                    # Browse and read variables from ComGate folder
                    root = self.client.nodes.objects
                    
                    # Try to find ComGate folder
                    children = await root.get_children()
                    for child in children:
                        name = await child.read_browse_name()
                        if "ComGate" in str(name):
                            variables = await child.get_children()
                            for var in variables[:20]:  # First 20
                                try:
                                    var_name = await var.read_browse_name()
                                    value = await var.read_value()
                                    self.received_data[str(var_name.Name)] = value
                                except:
                                    pass
                    
                    self.last_poll = datetime.now(timezone.utc)
                    self.error_message = None
                    
                except Exception as e:
                    self.error_message = str(e)
                    self.is_connected = False
                
                await asyncio.sleep(self.config.poll_interval_ms / 1000.0)
                
        except asyncio.CancelledError:
            pass

@api_router.post("/opcua-client/connect")
async def connect_opcua_client(
    config: OPCUAClientConfig,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Connect to a remote OPC UA server"""
    client_id = str(uuid.uuid4())
    client = OPCUAClientInstance(client_id, config)
    opcua_clients[client_id] = client
    
    connected = await client.connect()
    if connected:
        await client.start_polling()
    
    await log_audit(user, "opcua_client_connected", {
        "client_id": client_id,
        "endpoint": config.endpoint_url
    })
    
    return OPCUAClientInfo(
        id=client_id,
        endpoint_url=config.endpoint_url,
        security_mode=config.security_mode,
        is_connected=client.is_connected,
        poll_interval_ms=config.poll_interval_ms,
        created_at=client.created_at.isoformat(),
        error_message=client.error_message
    )

@api_router.post("/opcua-client/{client_id}/disconnect")
async def disconnect_opcua_client(
    client_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Disconnect an OPC UA client"""
    if client_id not in opcua_clients:
        raise HTTPException(status_code=404, detail="OPC UA client not found")
    
    await opcua_clients[client_id].stop_polling()
    del opcua_clients[client_id]
    
    await log_audit(user, "opcua_client_disconnected", {"client_id": client_id})
    return {"message": "OPC UA client disconnected"}

@api_router.get("/opcua-client/list")
async def list_opcua_clients(user: User = Depends(get_current_user)):
    """List all active OPC UA clients"""
    return [
        OPCUAClientInfo(
            id=client.id,
            endpoint_url=client.config.endpoint_url,
            security_mode=client.config.security_mode,
            is_connected=client.is_connected,
            poll_interval_ms=client.config.poll_interval_ms,
            created_at=client.created_at.isoformat(),
            last_poll=client.last_poll.isoformat() if client.last_poll else None,
            error_message=client.error_message
        )
        for client in opcua_clients.values()
    ]

@api_router.get("/opcua-client/{client_id}/data")
async def get_opcua_client_data(
    client_id: str,
    user: User = Depends(get_current_user)
):
    """Get data received from OPC UA server"""
    if client_id not in opcua_clients:
        raise HTTPException(status_code=404, detail="OPC UA client not found")
    
    client = opcua_clients[client_id]
    return {
        "is_connected": client.is_connected,
        "last_poll": client.last_poll.isoformat() if client.last_poll else None,
        "error_message": client.error_message,
        "variables": client.received_data
    }

# ==================== OPC DA SERVER & CLIENT ====================
# Note: OPC DA is Windows COM/DCOM based. This implementation provides
# a gateway-style interface that can work with OpenOPC2 on Windows.

opcda_servers = {}
opcda_clients = {}

class OPCDAServerConfig(BaseModel):
    server_name: str = "ComGate.OPC.DA.Server"
    prog_id: str = "ComGate.OPC.Simulation"
    num_tags: int = 100

class OPCDAServerInfo(BaseModel):
    id: str
    server_name: str
    prog_id: str
    is_running: bool
    simulation_enabled: bool
    created_at: str
    note: str = "OPC DA requires Windows COM. This is a simulation mode for development."

class OPCDAServerInstance:
    """OPC DA Server simulation (actual COM server requires Windows)"""
    
    def __init__(self, server_id: str, config: OPCDAServerConfig):
        self.id = server_id
        self.config = config
        self.is_running = False
        self.simulation_enabled = False
        self.simulation_task = None
        self.created_at = datetime.now(timezone.utc)
        self.tags = {f"Tag.{i:04d}": 0.0 for i in range(config.num_tags)}
    
    async def start(self):
        self.is_running = True
        logger.info(f"OPC DA Server {self.id} started (simulation mode)")
    
    async def stop(self):
        if self.simulation_task:
            self.simulation_task.cancel()
            try:
                await self.simulation_task
            except asyncio.CancelledError:
                pass
        self.is_running = False
    
    async def start_simulation(self, interval_ms: int = 1000, pattern: str = "sine"):
        if self.simulation_enabled:
            return
        
        self.simulation_enabled = True
        self.simulation_task = asyncio.create_task(self._simulation_loop(interval_ms, pattern))
    
    async def stop_simulation(self):
        self.simulation_enabled = False
        if self.simulation_task:
            self.simulation_task.cancel()
    
    async def _simulation_loop(self, interval_ms: int, pattern: str):
        start_time = datetime.now(timezone.utc)
        try:
            while self.simulation_enabled:
                elapsed = (datetime.now(timezone.utc) - start_time).total_seconds()
                
                for i, name in enumerate(self.tags.keys()):
                    phase_offset = i * (math.pi / 10)
                    
                    if pattern == "sine":
                        value = 50.0 + 50.0 * math.sin((2 * math.pi * elapsed / 60.0) + phase_offset)
                    elif pattern == "ramp":
                        cycle_pos = (elapsed % 60.0) / 60.0
                        value = 100.0 * cycle_pos
                    else:
                        import random
                        value = random.uniform(0, 100)
                    
                    self.tags[name] = value
                
                await asyncio.sleep(interval_ms / 1000.0)
        except asyncio.CancelledError:
            pass
    
    def get_tags(self) -> dict:
        return {k: v for k, v in list(self.tags.items())[:20]}

@api_router.post("/opcda-server/start")
async def start_opcda_server(
    config: OPCDAServerConfig,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Start an OPC DA server (simulation mode - actual COM server requires Windows)"""
    server_id = str(uuid.uuid4())
    server = OPCDAServerInstance(server_id, config)
    await server.start()
    opcda_servers[server_id] = server
    
    await log_audit(user, "opcda_server_started", {"server_id": server_id})
    
    return OPCDAServerInfo(
        id=server_id,
        server_name=config.server_name,
        prog_id=config.prog_id,
        is_running=True,
        simulation_enabled=False,
        created_at=server.created_at.isoformat()
    )

@api_router.post("/opcda-server/{server_id}/stop")
async def stop_opcda_server(
    server_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Stop an OPC DA server"""
    if server_id not in opcda_servers:
        raise HTTPException(status_code=404, detail="OPC DA server not found")
    
    await opcda_servers[server_id].stop()
    del opcda_servers[server_id]
    
    return {"message": "OPC DA server stopped"}

@api_router.get("/opcda-server/list")
async def list_opcda_servers(user: User = Depends(get_current_user)):
    """List all OPC DA servers"""
    return [
        OPCDAServerInfo(
            id=srv.id,
            server_name=srv.config.server_name,
            prog_id=srv.config.prog_id,
            is_running=srv.is_running,
            simulation_enabled=srv.simulation_enabled,
            created_at=srv.created_at.isoformat()
        )
        for srv in opcda_servers.values()
    ]

@api_router.get("/opcda-server/{server_id}/data")
async def get_opcda_server_data(
    server_id: str,
    user: User = Depends(get_current_user)
):
    """Get OPC DA server tag values"""
    if server_id not in opcda_servers:
        raise HTTPException(status_code=404, detail="OPC DA server not found")
    
    server = opcda_servers[server_id]
    return {
        "is_running": server.is_running,
        "simulation_enabled": server.simulation_enabled,
        "tags": server.get_tags()
    }

@api_router.post("/opcda-server/{server_id}/simulation/start")
async def start_opcda_simulation(
    server_id: str,
    config: SimulationConfig,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Start data simulation for OPC DA server"""
    if server_id not in opcda_servers:
        raise HTTPException(status_code=404, detail="OPC DA server not found")
    
    await opcda_servers[server_id].start_simulation(config.interval_ms, config.pattern)
    return {"message": "Simulation started"}

@api_router.post("/opcda-server/{server_id}/simulation/stop")
async def stop_opcda_simulation(
    server_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Stop data simulation for OPC DA server"""
    if server_id not in opcda_servers:
        raise HTTPException(status_code=404, detail="OPC DA server not found")
    
    await opcda_servers[server_id].stop_simulation()
    return {"message": "Simulation stopped"}

# OPC DA Client (simulation - actual connection requires Windows COM)
class OPCDAClientConfig(BaseModel):
    server_prog_id: str  # e.g., "Matrikon.OPC.Simulation"
    host: str = "localhost"
    poll_interval_ms: int = 1000

class OPCDAClientInfo(BaseModel):
    id: str
    server_prog_id: str
    host: str
    is_connected: bool
    poll_interval_ms: int
    created_at: str
    last_poll: Optional[str] = None
    error_message: Optional[str] = None
    note: str = "OPC DA requires Windows COM. Simulated connection for development."

class OPCDAClientInstance:
    """OPC DA Client (simulation mode - actual COM connection requires Windows)"""
    
    def __init__(self, client_id: str, config: OPCDAClientConfig):
        self.id = client_id
        self.config = config
        self.is_connected = False
        self.is_running = False
        self.task = None
        self.created_at = datetime.now(timezone.utc)
        self.last_poll = None
        self.error_message = None
        self.received_data = {}
    
    async def connect(self):
        # In simulation mode, we simulate a successful connection
        self.is_connected = True
        self.error_message = "Simulated connection (actual COM requires Windows)"
        logger.info(f"OPC DA Client {self.id} connected (simulation)")
        return True
    
    async def disconnect(self):
        self.is_connected = False
    
    async def start_polling(self):
        if self.is_running:
            return
        
        self.is_running = True
        self.task = asyncio.create_task(self._poll_loop())
    
    async def stop_polling(self):
        self.is_running = False
        if self.task:
            self.task.cancel()
            try:
                await self.task
            except asyncio.CancelledError:
                pass
    
    async def _poll_loop(self):
        """Simulate reading from OPC DA server"""
        try:
            while self.is_running:
                # Generate simulated data
                elapsed = (datetime.now(timezone.utc) - self.created_at).total_seconds()
                
                for i in range(20):
                    tag_name = f"Tag.{i:04d}"
                    phase_offset = i * (math.pi / 10)
                    value = 50.0 + 50.0 * math.sin((2 * math.pi * elapsed / 60.0) + phase_offset)
                    self.received_data[tag_name] = round(value, 2)
                
                self.last_poll = datetime.now(timezone.utc)
                await asyncio.sleep(self.config.poll_interval_ms / 1000.0)
                
        except asyncio.CancelledError:
            pass

@api_router.post("/opcda-client/connect")
async def connect_opcda_client(
    config: OPCDAClientConfig,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Connect to an OPC DA server (simulation mode)"""
    client_id = str(uuid.uuid4())
    client = OPCDAClientInstance(client_id, config)
    opcda_clients[client_id] = client
    
    await client.connect()
    await client.start_polling()
    
    await log_audit(user, "opcda_client_connected", {"client_id": client_id})
    
    return OPCDAClientInfo(
        id=client_id,
        server_prog_id=config.server_prog_id,
        host=config.host,
        is_connected=client.is_connected,
        poll_interval_ms=config.poll_interval_ms,
        created_at=client.created_at.isoformat(),
        error_message=client.error_message
    )

@api_router.post("/opcda-client/{client_id}/disconnect")
async def disconnect_opcda_client(
    client_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Disconnect an OPC DA client"""
    if client_id not in opcda_clients:
        raise HTTPException(status_code=404, detail="OPC DA client not found")
    
    await opcda_clients[client_id].stop_polling()
    del opcda_clients[client_id]
    
    return {"message": "OPC DA client disconnected"}

@api_router.get("/opcda-client/list")
async def list_opcda_clients(user: User = Depends(get_current_user)):
    """List all OPC DA clients"""
    return [
        OPCDAClientInfo(
            id=client.id,
            server_prog_id=client.config.server_prog_id,
            host=client.config.host,
            is_connected=client.is_connected,
            poll_interval_ms=client.config.poll_interval_ms,
            created_at=client.created_at.isoformat(),
            last_poll=client.last_poll.isoformat() if client.last_poll else None,
            error_message=client.error_message
        )
        for client in opcda_clients.values()
    ]

@api_router.get("/opcda-client/{client_id}/data")
async def get_opcda_client_data(
    client_id: str,
    user: User = Depends(get_current_user)
):
    """Get data received from OPC DA server"""
    if client_id not in opcda_clients:
        raise HTTPException(status_code=404, detail="OPC DA client not found")
    
    client = opcda_clients[client_id]
    return {
        "is_connected": client.is_connected,
        "last_poll": client.last_poll.isoformat() if client.last_poll else None,
        "error_message": client.error_message,
        "tags": client.received_data
    }

# ==================== POLLING ENGINE ====================
polling_tasks = {}

# Device connection pool for real Modbus polling
device_connections: Dict[str, Any] = {}

class RealModbusPoller:
    """Handles real Modbus TCP/UDP connections for a device"""
    
    def __init__(self, device: dict):
        self.device = device
        self.client = None
        self.is_connected = False
        self.last_error = None
        self.reconnect_attempts = 0
        self.max_reconnect_attempts = 3
    
    async def connect(self) -> bool:
        """Connect to the real Modbus device"""
        try:
            protocol = self.device.get('protocol', 'tcp')
            ip = self.device.get('ip_address')
            port = self.device.get('port', 502)
            
            if not ip:
                self.last_error = "No IP address configured"
                return False
            
            if protocol == 'tcp':
                self.client = AsyncModbusTcpClient(
                    host=ip,
                    port=port,
                    timeout=self.device.get('timeout_ms', 3000) / 1000.0
                )
            elif protocol == 'udp':
                self.client = AsyncModbusUdpClient(
                    host=ip,
                    port=port,
                    timeout=self.device.get('timeout_ms', 3000) / 1000.0
                )
            else:
                self.last_error = f"Unsupported protocol: {protocol}"
                return False
            
            connected = await self.client.connect()
            self.is_connected = connected
            if connected:
                self.last_error = None
                self.reconnect_attempts = 0
                logger.info(f"Connected to device {self.device['name']} at {ip}:{port}")
            else:
                self.last_error = f"Failed to connect to {ip}:{port}"
            return connected
        except Exception as e:
            self.last_error = str(e)
            self.is_connected = False
            logger.error(f"Connection error for {self.device['name']}: {e}")
            return False
    
    async def disconnect(self):
        """Disconnect from device"""
        if self.client:
            try:
                self.client.close()
            except:
                pass
        self.is_connected = False
        self.client = None
    
    async def ensure_connected(self) -> bool:
        """Ensure connection is established, reconnect if needed"""
        if self.is_connected and self.client:
            return True
        
        if self.reconnect_attempts >= self.max_reconnect_attempts:
            return False
        
        self.reconnect_attempts += 1
        return await self.connect()
    
    async def read_registers(self, obj_type: str, address: int, count: int, unit_id: int) -> tuple:
        """Read registers from device. Returns (values, error)"""
        if not await self.ensure_connected():
            return None, self.last_error or "Not connected"
        
        try:
            if obj_type == ObjectType.COIL.value:
                result = await self.client.read_coils(address, count, slave=unit_id)
            elif obj_type == ObjectType.DISCRETE_INPUT.value:
                result = await self.client.read_discrete_inputs(address, count, slave=unit_id)
            elif obj_type == ObjectType.INPUT_REGISTER.value:
                result = await self.client.read_input_registers(address, count, slave=unit_id)
            elif obj_type == ObjectType.HOLDING_REGISTER.value:
                result = await self.client.read_holding_registers(address, count, slave=unit_id)
            else:
                return None, f"Unknown object type: {obj_type}"
            
            if result.isError():
                self.is_connected = False
                return None, f"Modbus error: {result}"
            
            # Extract values based on type
            if obj_type in [ObjectType.COIL.value, ObjectType.DISCRETE_INPUT.value]:
                return list(result.bits[:count]), None
            else:
                return list(result.registers), None
                
        except Exception as e:
            self.is_connected = False
            self.last_error = str(e)
            return None, str(e)
    
    async def write_register(self, obj_type: str, address: int, value: Any, unit_id: int) -> tuple:
        """Write to device. Returns (success, error)"""
        if not await self.ensure_connected():
            return False, self.last_error or "Not connected"
        
        try:
            if obj_type == ObjectType.COIL.value:
                result = await self.client.write_coil(address, bool(value), slave=unit_id)
            elif obj_type == ObjectType.HOLDING_REGISTER.value:
                result = await self.client.write_register(address, int(value) & 0xFFFF, slave=unit_id)
            else:
                return False, f"Cannot write to {obj_type}"
            
            if result.isError():
                return False, f"Modbus write error: {result}"
            
            return True, None
        except Exception as e:
            self.last_error = str(e)
            return False, str(e)
    
    async def write_registers(self, obj_type: str, address: int, values: List[int], unit_id: int) -> tuple:
        """Write multiple registers. Returns (success, error)"""
        if not await self.ensure_connected():
            return False, self.last_error or "Not connected"
        
        try:
            if obj_type == ObjectType.COIL.value:
                result = await self.client.write_coils(address, [bool(v) for v in values], slave=unit_id)
            elif obj_type == ObjectType.HOLDING_REGISTER.value:
                result = await self.client.write_registers(address, [int(v) & 0xFFFF for v in values], slave=unit_id)
            else:
                return False, f"Cannot write to {obj_type}"
            
            if result.isError():
                return False, f"Modbus write error: {result}"
            
            return True, None
        except Exception as e:
            self.last_error = str(e)
            return False, str(e)


def decode_modbus_value(raw_registers: List[int], data_type: str, endian: str = 'ABCD') -> Any:
    """Decode raw Modbus registers to actual value based on data type and endianness"""
    if not raw_registers:
        return None
    
    if data_type == 'bool':
        return bool(raw_registers[0])
    elif data_type == 'uint16':
        return raw_registers[0]
    elif data_type == 'int16':
        val = raw_registers[0]
        return val if val < 32768 else val - 65536
    elif data_type in ['uint32', 'int32', 'float32']:
        if len(raw_registers) < 2:
            return raw_registers[0] if raw_registers else 0
        
        # Get bytes based on endianness
        if endian == 'ABCD':  # Big-endian
            bytes_val = struct.pack('>HH', raw_registers[0], raw_registers[1])
        elif endian == 'CDAB':  # Little-endian word swap
            bytes_val = struct.pack('>HH', raw_registers[1], raw_registers[0])
        elif endian == 'BADC':  # Big-endian byte swap
            bytes_val = struct.pack('<HH', raw_registers[0], raw_registers[1])
        elif endian == 'DCBA':  # Little-endian
            bytes_val = struct.pack('<HH', raw_registers[1], raw_registers[0])
        else:
            bytes_val = struct.pack('>HH', raw_registers[0], raw_registers[1])
        
        if data_type == 'uint32':
            return struct.unpack('>I', bytes_val)[0]
        elif data_type == 'int32':
            return struct.unpack('>i', bytes_val)[0]
        elif data_type == 'float32':
            return struct.unpack('>f', bytes_val)[0]
    elif data_type == 'float64':
        if len(raw_registers) < 4:
            return 0.0
        # Combine 4 registers for 64-bit float
        if endian == 'ABCD':
            bytes_val = struct.pack('>HHHH', *raw_registers[:4])
        else:
            bytes_val = struct.pack('>HHHH', *raw_registers[:4])
        return struct.unpack('>d', bytes_val)[0]
    
    return raw_registers[0] if raw_registers else 0


def encode_modbus_value(value: Any, data_type: str, endian: str = 'ABCD') -> List[int]:
    """Encode a value to raw Modbus registers"""
    if data_type == 'bool':
        return [1 if value else 0]
    elif data_type == 'uint16':
        return [int(value) & 0xFFFF]
    elif data_type == 'int16':
        val = int(value)
        if val < 0:
            val = val + 65536
        return [val & 0xFFFF]
    elif data_type in ['uint32', 'int32', 'float32']:
        if data_type == 'uint32':
            bytes_val = struct.pack('>I', int(value))
        elif data_type == 'int32':
            bytes_val = struct.pack('>i', int(value))
        else:  # float32
            bytes_val = struct.pack('>f', float(value))
        
        regs = struct.unpack('>HH', bytes_val)
        
        if endian == 'ABCD':
            return list(regs)
        elif endian == 'CDAB':
            return [regs[1], regs[0]]
        elif endian == 'BADC':
            return list(struct.unpack('<HH', bytes_val))
        elif endian == 'DCBA':
            r = struct.unpack('<HH', bytes_val)
            return [r[1], r[0]]
        return list(regs)
    
    return [int(value) & 0xFFFF]


def get_register_count(data_type: str) -> int:
    """Get number of registers needed for a data type"""
    if data_type in ['bool', 'int16', 'uint16']:
        return 1
    elif data_type in ['int32', 'uint32', 'float32']:
        return 2
    elif data_type == 'float64':
        return 4
    return 1


class PollingEngine:
    """Enhanced polling engine that supports both real Modbus and simulation"""
    
    def __init__(self, project_id: str):
        self.project_id = project_id
        self.is_running = False
        self.task = None
        self.device_pollers: Dict[str, RealModbusPoller] = {}
        self.use_simulation = False  # Will be set based on device connectivity
    
    async def start(self):
        self.is_running = True
        self.task = asyncio.create_task(self.poll_loop())
        logger.info(f"Polling started for project {self.project_id}")
    
    async def stop(self):
        self.is_running = False
        if self.task:
            self.task.cancel()
            try:
                await self.task
            except asyncio.CancelledError:
                pass
        
        # Disconnect all device pollers
        for poller in self.device_pollers.values():
            await poller.disconnect()
        self.device_pollers.clear()
        logger.info(f"Polling stopped for project {self.project_id}")
    
    def get_or_create_poller(self, device: dict) -> RealModbusPoller:
        """Get existing poller or create new one for device"""
        device_id = device['id']
        if device_id not in self.device_pollers:
            self.device_pollers[device_id] = RealModbusPoller(device)
        return self.device_pollers[device_id]
    
    async def poll_loop(self):
        """Main polling loop"""
        while self.is_running:
            try:
                devices = await db.devices.find(
                    {"project_id": self.project_id, "is_enabled": True}, 
                    {"_id": 0}
                ).to_list(100)
                
                for device in devices:
                    if not self.is_running:
                        break
                    
                    # Get tags for this device
                    tags = await db.tags.find({
                        "device_id": device['id'],
                        "is_forced": {"$ne": True}  # Skip forced tags
                    }, {"_id": 0}).to_list(5000)
                    
                    if not tags:
                        continue
                    
                    # Try real Modbus first
                    if device.get('ip_address'):
                        await self.poll_device_real(device, tags)
                    else:
                        # Fall back to simulation
                        await self.poll_device_simulated(device, tags)
                
                # Small delay between poll cycles
                await asyncio.sleep(0.5)
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Polling error: {e}")
                await asyncio.sleep(2)
    
    async def poll_device_real(self, device: dict, tags: List[dict]):
        """Poll tags from a real Modbus device"""
        poller = self.get_or_create_poller(device)
        unit_id = device.get('unit_id', 1)
        device_endian = device.get('default_endian', 'ABCD')
        
        # Group tags by object type for efficient block reads
        tags_by_type: Dict[str, List[dict]] = {}
        for tag in tags:
            obj_type = tag.get('object_type', 'holding_register')
            if obj_type not in tags_by_type:
                tags_by_type[obj_type] = []
            tags_by_type[obj_type].append(tag)
        
        # Poll each object type
        for obj_type, type_tags in tags_by_type.items():
            if not type_tags:
                continue
            
            # Sort by address and find optimal read blocks
            type_tags.sort(key=lambda t: t.get('address', 0))
            
            # Create read blocks (max block size from device config)
            max_block = device.get('max_block_size', 120)
            blocks = self.create_read_blocks(type_tags, max_block)
            
            for block_start, block_count, block_tags in blocks:
                start_time = datetime.now(timezone.utc)
                
                # Read from device
                values, error = await poller.read_registers(obj_type, block_start, block_count, unit_id)
                
                end_time = datetime.now(timezone.utc)
                rtt = (end_time - start_time).total_seconds() * 1000
                
                if error:
                    # Log error and update tags as bad quality
                    await self.log_traffic(
                        device, obj_type, block_start, block_count,
                        None, rtt, TrafficStatus.ERROR, error
                    )
                    
                    for tag in block_tags:
                        await self.update_tag_error(tag, error)
                    
                    await self.update_device_status(device, "error", error)
                    
                    # Fall back to simulation for this device
                    await self.poll_device_simulated(device, tags)
                    return
                else:
                    # Decode and update each tag
                    for tag in block_tags:
                        tag_address = tag.get('address', 0)
                        offset_in_block = tag_address - block_start
                        reg_count = get_register_count(tag.get('data_type', 'uint16'))
                        
                        if offset_in_block >= 0 and offset_in_block + reg_count <= len(values):
                            raw_vals = values[offset_in_block:offset_in_block + reg_count]
                            endian = tag.get('endian') or device_endian
                            
                            # Decode value
                            decoded = decode_modbus_value(raw_vals, tag.get('data_type', 'uint16'), endian)
                            
                            # Apply scale and offset
                            if decoded is not None and tag.get('data_type') not in ['bool']:
                                scale = tag.get('scale', 1.0) or 1.0
                                offset = tag.get('offset', 0.0) or 0.0
                                decoded = decoded * scale + offset
                            
                            await self.update_tag_value(tag, decoded, TagQuality.GOOD.value)
                    
                    # Log successful traffic
                    await self.log_traffic(
                        device, obj_type, block_start, block_count,
                        f"Read {len(block_tags)} tags", rtt, TrafficStatus.OK, None
                    )
                    
                    await self.update_device_status(device, "online", None)
    
    async def poll_device_simulated(self, device: dict, tags: List[dict]):
        """Poll using local simulator or generate simulated values"""
        import random
        import math
        
        start_time = datetime.now(timezone.utc)
        
        # Check for active simulator sessions or modbus servers
        sim_source = None
        
        # First check modbus_servers (run as slave)
        for sid, srv in modbus_servers.items():
            if srv.is_running:
                sim_source = ('server', srv)
                break
        
        # Then check simulator_sessions (local simulator)
        if not sim_source:
            for sid, sim in simulator_sessions.items():
                if sim.is_running:
                    sim_source = ('simulator', sim)
                    break
        
        for tag in tags:
            obj_type = tag.get('object_type', 'holding_register')
            address = tag.get('address', 0)
            data_type = tag.get('data_type', 'uint16')
            
            value = None
            
            if sim_source:
                source_type, source = sim_source
                try:
                    if source_type == 'server':
                        # Read from modbus server datastore
                        if obj_type == ObjectType.COIL.value:
                            vals = source.read_registers("coil", address, 1)
                            value = bool(vals[0]) if vals else False
                        elif obj_type == ObjectType.DISCRETE_INPUT.value:
                            vals = source.read_registers("discrete_input", address, 1)
                            value = bool(vals[0]) if vals else False
                        elif obj_type == ObjectType.INPUT_REGISTER.value:
                            vals = source.read_registers("input_register", address, get_register_count(data_type))
                            value = decode_modbus_value(vals, data_type, tag.get('endian', 'ABCD'))
                        else:
                            vals = source.read_registers("holding_register", address, get_register_count(data_type))
                            value = decode_modbus_value(vals, data_type, tag.get('endian', 'ABCD'))
                    else:
                        # Read from simulator
                        if obj_type == ObjectType.COIL.value:
                            vals = source.read_coils(address, 1)
                            value = vals[0] if vals else False
                        elif obj_type == ObjectType.DISCRETE_INPUT.value:
                            vals = source.read_discrete_inputs(address, 1)
                            value = vals[0] if vals else False
                        elif obj_type == ObjectType.INPUT_REGISTER.value:
                            vals = source.read_input_registers(address, get_register_count(data_type))
                            value = decode_modbus_value(vals, data_type, tag.get('endian', 'ABCD')) if vals else 0
                        else:
                            vals = source.read_holding_registers(address, get_register_count(data_type))
                            value = decode_modbus_value(vals, data_type, tag.get('endian', 'ABCD')) if vals else 0
                except Exception as e:
                    logger.warning(f"Simulator read error: {e}")
                    value = None
            
            # Generate simulated value if no source available
            if value is None:
                if data_type == 'bool':
                    value = random.choice([True, False])
                else:
                    # Use sine wave for realistic simulation
                    elapsed = (datetime.now(timezone.utc) - datetime(2024, 1, 1, tzinfo=timezone.utc)).total_seconds()
                    min_val = tag.get('min_value') or 0
                    max_val = tag.get('max_value') or 100
                    amplitude = (max_val - min_val) / 2
                    offset = min_val + amplitude
                    # Add address-based phase offset for variety
                    phase = address * 0.1
                    value = offset + amplitude * math.sin(elapsed * 0.1 + phase)
                    value = round(value, 2)
            
            # Apply scale and offset for non-bool types
            if value is not None and data_type not in ['bool']:
                scale = tag.get('scale', 1.0) or 1.0
                offset = tag.get('offset', 0.0) or 0.0
                value = value * scale + offset
            
            await self.update_tag_value(tag, value, TagQuality.GOOD.value)
        
        end_time = datetime.now(timezone.utc)
        rtt = (end_time - start_time).total_seconds() * 1000
        
        # Log simulated traffic
        await self.log_traffic(
            device, "simulation", 0, len(tags),
            f"Simulated {len(tags)} tags", rtt, TrafficStatus.OK, None
        )
        
        await self.update_device_status(device, "simulated", None)
    
    def create_read_blocks(self, tags: List[dict], max_block_size: int) -> List[tuple]:
        """Create optimal read blocks from sorted tags"""
        if not tags:
            return []
        
        blocks = []
        current_start = tags[0].get('address', 0)
        current_tags = [tags[0]]
        
        for tag in tags[1:]:
            tag_addr = tag.get('address', 0)
            reg_count = get_register_count(tag.get('data_type', 'uint16'))
            current_end = current_start + sum(
                get_register_count(t.get('data_type', 'uint16')) 
                for t in current_tags
            )
            
            # Check if we can extend the current block
            gap = tag_addr - current_end
            potential_block_size = tag_addr - current_start + reg_count
            
            if gap <= 10 and potential_block_size <= max_block_size:
                # Extend current block
                current_tags.append(tag)
            else:
                # Save current block and start new one
                block_count = max(
                    t.get('address', 0) + get_register_count(t.get('data_type', 'uint16')) 
                    for t in current_tags
                ) - current_start
                blocks.append((current_start, block_count, current_tags))
                
                current_start = tag_addr
                current_tags = [tag]
        
        # Don't forget the last block
        if current_tags:
            block_count = max(
                t.get('address', 0) + get_register_count(t.get('data_type', 'uint16')) 
                for t in current_tags
            ) - current_start
            blocks.append((current_start, block_count, current_tags))
        
        return blocks
    
    async def update_tag_value(self, tag: dict, value: Any, quality: str):
        """Update tag value in database and broadcast via WebSocket"""
        now = datetime.now(timezone.utc)
        
        await db.tags.update_one(
            {"id": tag['id']},
            {"$set": {
                "current_value": value,
                "quality": quality,
                "last_update": now.isoformat(),
                "error_message": None
            }}
        )
        
        # Broadcast update via WebSocket
        await broadcast_tag_update(self.project_id, tag['id'], value, quality)
    
    async def update_tag_error(self, tag: dict, error: str):
        """Update tag with error status"""
        await db.tags.update_one(
            {"id": tag['id']},
            {"$set": {
                "quality": TagQuality.BAD.value,
                "error_message": error
            }}
        )
        
        await broadcast_tag_update(self.project_id, tag['id'], tag.get('current_value'), TagQuality.BAD.value)
    
    async def update_device_status(self, device: dict, status: str, error: Optional[str]):
        """Update device status in database"""
        update = {"status": status, "last_poll": datetime.now(timezone.utc).isoformat()}
        if error:
            update["$inc"] = {"error_count": 1}
        else:
            update["$inc"] = {"success_count": 1}
        
        # Separate $set and $inc
        set_update = {"status": status, "last_poll": datetime.now(timezone.utc).isoformat()}
        inc_update = {"error_count": 1} if error else {"success_count": 1}
        
        await db.devices.update_one(
            {"id": device['id']},
            {"$set": set_update, "$inc": inc_update}
        )
    
    async def log_traffic(self, device: dict, obj_type: str, start_addr: int, count: int,
                          response: Optional[str], rtt: float, status: TrafficStatus, error: Optional[str]):
        """Log traffic to database"""
        fc_map = {
            'coil': 1, 'discrete_input': 2, 
            'input_register': 4, 'holding_register': 3,
            'simulation': 0
        }
        
        traffic = TrafficLog(
            project_id=self.project_id,
            device_id=device['id'],
            device_name=device['name'],
            protocol=device.get('protocol', 'tcp'),
            function_code=fc_map.get(obj_type, 3),
            request_summary=f"Read {obj_type} @ {start_addr} x{count}",
            response_summary=response,
            round_trip_ms=rtt,
            status=status,
            error_details=error
        )
        doc = traffic.model_dump()
        doc['timestamp'] = doc['timestamp'].isoformat()
        await db.traffic_logs.insert_one(doc)

@api_router.post("/projects/{project_id}/polling/start")
async def start_polling(project_id: str, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER, UserRole.OPERATOR]))):
    if project_id in polling_tasks:
        raise HTTPException(status_code=400, detail="Polling already running")
    
    engine = PollingEngine(project_id)
    await engine.start()
    polling_tasks[project_id] = engine
    
    await log_audit(user, "polling_started", {"project_id": project_id})
    return {"message": "Polling started"}

@api_router.post("/projects/{project_id}/polling/stop")
async def stop_polling(project_id: str, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER, UserRole.OPERATOR]))):
    if project_id not in polling_tasks:
        raise HTTPException(status_code=400, detail="Polling not running")
    
    await polling_tasks[project_id].stop()
    del polling_tasks[project_id]
    
    await log_audit(user, "polling_stopped", {"project_id": project_id})
    return {"message": "Polling stopped"}

@api_router.get("/projects/{project_id}/polling/status")
async def get_polling_status(project_id: str, user: User = Depends(get_current_user)):
    is_running = project_id in polling_tasks and polling_tasks[project_id].is_running
    return {"is_running": is_running}

class TagWriteRequest(BaseModel):
    value: Any

class TagForceRequest(BaseModel):
    value: Any
    force: bool = False  # If True, bypasses read-only check

# ==================== WRITE ENGINE ====================
async def write_to_real_device(device: dict, tag: dict, raw_value: Any) -> tuple:
    """Write to real Modbus device. Returns (success, error)"""
    try:
        protocol = device.get('protocol', 'tcp')
        ip = device.get('ip_address')
        port = device.get('port', 502)
        unit_id = device.get('unit_id', 1)
        
        if not ip:
            return False, "No IP address configured"
        
        # Create temporary client for write
        if protocol == 'tcp':
            client = AsyncModbusTcpClient(host=ip, port=port, timeout=device.get('timeout_ms', 3000) / 1000.0)
        elif protocol == 'udp':
            client = AsyncModbusUdpClient(host=ip, port=port, timeout=device.get('timeout_ms', 3000) / 1000.0)
        else:
            return False, f"Unsupported protocol: {protocol}"
        
        connected = await client.connect()
        if not connected:
            return False, f"Failed to connect to {ip}:{port}"
        
        try:
            obj_type = tag.get('object_type')
            address = tag.get('address', 0)
            data_type = tag.get('data_type', 'uint16')
            endian = tag.get('endian') or device.get('default_endian', 'ABCD')
            
            if obj_type == ObjectType.COIL.value:
                result = await client.write_coil(address, bool(raw_value), slave=unit_id)
            elif obj_type == ObjectType.HOLDING_REGISTER.value:
                # Encode value to registers
                registers = encode_modbus_value(raw_value, data_type, endian)
                if len(registers) == 1:
                    result = await client.write_register(address, registers[0], slave=unit_id)
                else:
                    result = await client.write_registers(address, registers, slave=unit_id)
            else:
                return False, f"Cannot write to {obj_type}"
            
            if result.isError():
                return False, f"Modbus write error: {result}"
            
            return True, None
        finally:
            client.close()
            
    except Exception as e:
        return False, str(e)


async def write_to_simulator(tag: dict, raw_value: Any) -> bool:
    """Write to local simulator or modbus server"""
    obj_type = tag.get('object_type')
    address = tag.get('address', 0)
    
    # Try modbus_servers first
    for sid, srv in modbus_servers.items():
        if srv.is_running:
            if obj_type == ObjectType.COIL.value:
                srv.write_registers("coil", address, [bool(raw_value)])
            elif obj_type == ObjectType.HOLDING_REGISTER.value:
                srv.write_registers("holding_register", address, [int(raw_value) & 0xFFFF])
            return True
    
    # Then try simulator_sessions
    for sid, sim in simulator_sessions.items():
        if sim.is_running:
            if obj_type == ObjectType.COIL.value:
                sim.write_coil(address, bool(raw_value))
            elif obj_type == ObjectType.HOLDING_REGISTER.value:
                sim.write_register(address, int(raw_value))
            return True
    
    return False


@api_router.post("/tags/{tag_id}/write")
async def write_tag(
    tag_id: str,
    request: TagWriteRequest,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER, UserRole.OPERATOR]))
):
    value = request.value
    tag = await db.tags.find_one({"id": tag_id}, {"_id": 0})
    if not tag:
        raise HTTPException(status_code=404, detail="Tag not found")
    
    if tag.get('permission') == TagPermission.READ.value:
        raise HTTPException(status_code=403, detail="Tag is read-only")
    
    # Validate bounds
    if tag.get('min_value') is not None and value < tag['min_value']:
        raise HTTPException(status_code=400, detail=f"Value below minimum ({tag['min_value']})")
    if tag.get('max_value') is not None and value > tag['max_value']:
        raise HTTPException(status_code=400, detail=f"Value above maximum ({tag['max_value']})")
    
    # Get device
    device = await db.devices.find_one({"id": tag['device_id']}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    
    # Calculate raw value (reverse scale/offset)
    scale = tag.get('scale', 1.0) or 1.0
    offset = tag.get('offset', 0.0) or 0.0
    raw_value = (value - offset) / scale if scale != 0 else value
    
    start_time = datetime.now(timezone.utc)
    write_error = None
    write_success = False
    
    # Try real device first if IP is configured
    if device.get('ip_address'):
        write_success, write_error = await write_to_real_device(device, tag, raw_value)
    
    # Fall back to simulator
    if not write_success:
        sim_success = await write_to_simulator(tag, raw_value)
        if sim_success:
            write_success = True
            write_error = None
    
    end_time = datetime.now(timezone.utc)
    rtt = (end_time - start_time).total_seconds() * 1000
    
    if not write_success and write_error:
        # Log failed traffic
        traffic = TrafficLog(
            project_id=tag['project_id'],
            device_id=device['id'],
            device_name=device['name'],
            protocol=device.get('protocol', 'tcp'),
            function_code=6 if tag.get('object_type') == ObjectType.HOLDING_REGISTER.value else 5,
            request_summary=f"Write {tag['name']} = {value}",
            response_summary=None,
            round_trip_ms=rtt,
            status=TrafficStatus.ERROR,
            error_details=write_error
        )
        doc = traffic.model_dump()
        doc['timestamp'] = doc['timestamp'].isoformat()
        await db.traffic_logs.insert_one(doc)
        
        raise HTTPException(status_code=500, detail=f"Write failed: {write_error}")
    
    # Update tag value
    await db.tags.update_one(
        {"id": tag_id},
        {"$set": {
            "current_value": value,
            "last_update": end_time.isoformat()
        }}
    )
    
    # Broadcast update
    await broadcast_tag_update(tag['project_id'], tag_id, value, TagQuality.GOOD.value)
    
    # Log audit
    await log_audit(user, "tag_write", {
        "tag_id": tag_id,
        "tag_name": tag.get('name'),
        "value": value
    })
    
    # Log successful traffic
    traffic = TrafficLog(
        project_id=tag['project_id'],
        device_id=device['id'],
        device_name=device['name'],
        protocol=device.get('protocol', 'tcp'),
        function_code=6 if tag.get('object_type') == ObjectType.HOLDING_REGISTER.value else 5,
        request_summary=f"Write {tag['name']} = {value}",
        response_summary="OK",
        round_trip_ms=rtt,
        status=TrafficStatus.OK
    )
    doc = traffic.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    await db.traffic_logs.insert_one(doc)
    
    return {"message": "Write successful", "value": value}

# Force value endpoint - restricted to Admin/Engineer roles
@api_router.post("/tags/{tag_id}/force")
async def force_tag_value(
    tag_id: str,
    request: TagForceRequest,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Force a value to a tag, bypassing read-only restrictions. Requires Admin or Engineer role."""
    value = request.value
    tag = await db.tags.find_one({"id": tag_id}, {"_id": 0})
    if not tag:
        raise HTTPException(status_code=404, detail="Tag not found")
    
    # Write to simulator if available
    for sid, sim in simulator_sessions.items():
        if sim.is_running:
            obj_type = tag.get('object_type')
            address = tag.get('address', 0)
            
            if obj_type == ObjectType.COIL.value:
                sim.write_coil(address, bool(value))
            elif obj_type == ObjectType.DISCRETE_INPUT.value:
                if 0 <= address < len(sim.discrete_inputs):
                    sim.discrete_inputs[address] = bool(value)
            elif obj_type == ObjectType.INPUT_REGISTER.value:
                if 0 <= address < len(sim.input_registers):
                    sim.input_registers[address] = int(value) & 0xFFFF
            elif obj_type == ObjectType.HOLDING_REGISTER.value:
                raw_value = (value - tag.get('offset', 0)) / tag.get('scale', 1)
                sim.write_register(address, int(raw_value))
            break
    
    # Update tag value with forced flag
    await db.tags.update_one(
        {"id": tag_id},
        {"$set": {
            "current_value": value,
            "quality": TagQuality.GOOD.value,
            "last_update": datetime.now(timezone.utc).isoformat(),
            "is_forced": True
        }}
    )
    
    # Log audit
    await log_audit(user, "tag_forced", {
        "tag_id": tag_id,
        "tag_name": tag.get('name'),
        "value": value
    })
    
    return {"message": "Value forced successfully", "value": value}

# Release forced value
@api_router.post("/tags/{tag_id}/release")
async def release_tag_force(
    tag_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))
):
    """Release a forced value, allowing normal polling to resume."""
    tag = await db.tags.find_one({"id": tag_id}, {"_id": 0})
    if not tag:
        raise HTTPException(status_code=404, detail="Tag not found")
    
    await db.tags.update_one(
        {"id": tag_id},
        {"$set": {"is_forced": False}}
    )
    
    await log_audit(user, "tag_released", {"tag_id": tag_id, "tag_name": tag.get('name')})
    return {"message": "Force released"}

@api_router.post("/projects/{project_id}/batch-write")
async def batch_write(
    project_id: str,
    request: BatchWriteRequest,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER, UserRole.OPERATOR]))
):
    results = []
    for write in request.writes:
        try:
            write_request = TagWriteRequest(value=write.value)
            await write_tag(write.tag_id, write_request, user)
            results.append({"tag_id": write.tag_id, "success": True})
        except Exception as e:
            results.append({"tag_id": write.tag_id, "success": False, "error": str(e)})
    
    return {"results": results}

# ==================== REPORTS ====================
@api_router.get("/projects/{project_id}/reports/summary")
async def get_project_summary(project_id: str, user: User = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    devices = await db.devices.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    tags = await db.tags.find({"project_id": project_id}, {"_id": 0}).to_list(50000)
    traffic = await db.traffic_logs.find({"project_id": project_id}, {"_id": 0}).to_list(10000)
    
    # Calculate stats
    total_tags = len(tags)
    read_tags = len([t for t in tags if t.get('permission') in ['R', 'RW']])
    write_tags = len([t for t in tags if t.get('permission') in ['W', 'RW']])
    
    online_devices = len([d for d in devices if d.get('status') == 'online'])
    offline_devices = len(devices) - online_devices
    
    ok_traffic = len([t for t in traffic if t.get('status') == 'ok'])
    error_traffic = len(traffic) - ok_traffic
    
    rtts = [t.get('round_trip_ms', 0) for t in traffic if t.get('round_trip_ms')]
    avg_rtt = sum(rtts) / len(rtts) if rtts else 0
    max_rtt = max(rtts) if rtts else 0
    
    return {
        "project": project,
        "devices": {
            "total": len(devices),
            "online": online_devices,
            "offline": offline_devices,
            "by_protocol": {
                "tcp": len([d for d in devices if d.get('protocol') == 'tcp']),
                "udp": len([d for d in devices if d.get('protocol') == 'udp']),
                "rtu": len([d for d in devices if d.get('protocol') == 'rtu'])
            }
        },
        "tags": {
            "total": total_tags,
            "read": read_tags,
            "write": write_tags,
            "by_quality": {
                "good": len([t for t in tags if t.get('quality') == 'good']),
                "bad": len([t for t in tags if t.get('quality') == 'bad']),
                "uncertain": len([t for t in tags if t.get('quality') == 'uncertain'])
            }
        },
        "traffic": {
            "total": len(traffic),
            "success": ok_traffic,
            "errors": error_traffic,
            "error_rate": (error_traffic / len(traffic) * 100) if traffic else 0,
            "avg_rtt_ms": round(avg_rtt, 2),
            "max_rtt_ms": round(max_rtt, 2)
        }
    }

@api_router.get("/projects/{project_id}/reports/export")
async def export_report(project_id: str, user: User = Depends(get_current_user)):
    from fastapi.responses import StreamingResponse
    
    summary = await get_project_summary(project_id, user)
    
    output = BytesIO()
    workbook = Workbook()
    
    # Summary sheet
    ws = workbook.active
    ws.title = "Summary"
    ws.append(["Project Report"])
    ws.append(["Generated", datetime.now(timezone.utc).isoformat()])
    ws.append([])
    ws.append(["Project Name", summary['project'].get('name', '')])
    ws.append(["Description", summary['project'].get('description', '')])
    ws.append([])
    ws.append(["Devices"])
    ws.append(["Total", summary['devices']['total']])
    ws.append(["Online", summary['devices']['online']])
    ws.append(["Offline", summary['devices']['offline']])
    ws.append([])
    ws.append(["Tags"])
    ws.append(["Total", summary['tags']['total']])
    ws.append(["Read", summary['tags']['read']])
    ws.append(["Write", summary['tags']['write']])
    ws.append([])
    ws.append(["Performance"])
    ws.append(["Total Transactions", summary['traffic']['total']])
    ws.append(["Success", summary['traffic']['success']])
    ws.append(["Errors", summary['traffic']['errors']])
    ws.append(["Error Rate (%)", summary['traffic']['error_rate']])
    ws.append(["Avg RTT (ms)", summary['traffic']['avg_rtt_ms']])
    ws.append(["Max RTT (ms)", summary['traffic']['max_rtt_ms']])
    
    # Devices sheet
    ws2 = workbook.create_sheet("Devices")
    ws2.append(["Name", "Protocol", "IP/Port", "Unit ID", "Status", "Success Count", "Error Count"])
    devices = await db.devices.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    for d in devices:
        ws2.append([
            d.get('name', ''),
            d.get('protocol', ''),
            f"{d.get('ip_address', '')}:{d.get('port', '')}" if d.get('ip_address') else d.get('com_port', ''),
            d.get('unit_id', ''),
            d.get('status', ''),
            d.get('success_count', 0),
            d.get('error_count', 0)
        ])
    
    # Tags sheet
    ws3 = workbook.create_sheet("Tags")
    ws3.append(["Name", "Device", "Object Type", "Address", "Data Type", "Permission", "Quality", "Current Value"])
    tags = await db.tags.find({"project_id": project_id}, {"_id": 0}).to_list(50000)
    device_map = {d['id']: d['name'] for d in devices}
    for t in tags:
        ws3.append([
            t.get('name', ''),
            device_map.get(t.get('device_id', ''), ''),
            t.get('object_type', ''),
            t.get('address', ''),
            t.get('data_type', ''),
            t.get('permission', ''),
            t.get('quality', ''),
            t.get('current_value', '')
        ])
    
    workbook.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=report_{project_id}.xlsx"}
    )

@api_router.get("/projects/{project_id}/reports/export-pdf")
async def export_report_pdf(project_id: str, user: User = Depends(get_current_user)):
    """Export project report as PDF"""
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.units import inch
    
    summary = await get_project_summary(project_id, user)
    project = summary['project']
    
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=24, spaceAfter=30, textColor=colors.HexColor('#1e40af'))
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14, spaceBefore=20, spaceAfter=10, textColor=colors.HexColor('#1e40af'))
    normal_style = styles['Normal']
    
    elements = []
    
    # Title
    elements.append(Paragraph(f"FAT Report: {project.get('name', 'Project')}", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}", normal_style))
    elements.append(Spacer(1, 20))
    
    # Project Info
    elements.append(Paragraph("Project Information", heading_style))
    project_data = [
        ["Project Name", project.get('name', '')],
        ["Description", project.get('description', 'N/A')],
        ["Owner", project.get('owner_id', 'N/A')],
    ]
    project_table = Table(project_data, colWidths=[2*inch, 4*inch])
    project_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
    ]))
    elements.append(project_table)
    elements.append(Spacer(1, 20))
    
    # Device Summary
    elements.append(Paragraph("Device Summary", heading_style))
    device_data = [
        ["Metric", "Value"],
        ["Total Devices", str(summary['devices']['total'])],
        ["Online", str(summary['devices']['online'])],
        ["Offline", str(summary['devices']['offline'])],
        ["TCP Devices", str(summary['devices']['by_protocol']['tcp'])],
        ["UDP Devices", str(summary['devices']['by_protocol']['udp'])],
        ["RTU Devices", str(summary['devices']['by_protocol']['rtu'])],
    ]
    device_table = Table(device_data, colWidths=[3*inch, 2*inch])
    device_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f3f4f6')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
    ]))
    elements.append(device_table)
    elements.append(Spacer(1, 20))
    
    # Tag Summary
    elements.append(Paragraph("Tag Summary", heading_style))
    tag_data = [
        ["Metric", "Value"],
        ["Total Tags", str(summary['tags']['total'])],
        ["Read Tags", str(summary['tags']['read'])],
        ["Write Tags", str(summary['tags']['write'])],
        ["Good Quality", str(summary['tags']['by_quality']['good'])],
        ["Bad Quality", str(summary['tags']['by_quality']['bad'])],
        ["Uncertain", str(summary['tags']['by_quality']['uncertain'])],
    ]
    tag_table = Table(tag_data, colWidths=[3*inch, 2*inch])
    tag_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f3f4f6')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
    ]))
    elements.append(tag_table)
    elements.append(Spacer(1, 20))
    
    # Performance Summary
    elements.append(Paragraph("Performance Summary", heading_style))
    perf_data = [
        ["Metric", "Value"],
        ["Total Transactions", str(summary['traffic']['total'])],
        ["Successful", str(summary['traffic']['success'])],
        ["Errors", str(summary['traffic']['errors'])],
        ["Error Rate", f"{summary['traffic']['error_rate']:.2f}%"],
        ["Avg RTT", f"{summary['traffic']['avg_rtt_ms']:.2f} ms"],
        ["Max RTT", f"{summary['traffic']['max_rtt_ms']:.2f} ms"],
    ]
    perf_table = Table(perf_data, colWidths=[3*inch, 2*inch])
    perf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f3f4f6')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
    ]))
    elements.append(perf_table)
    elements.append(Spacer(1, 30))
    
    # FAT Checklist
    elements.append(Paragraph("FAT Checklist", heading_style))
    checklist_items = [
        ("Device Configuration Verified", summary['devices']['total'] > 0),
        ("Tag Mapping Complete", summary['tags']['total'] > 0),
        ("Communication Test Passed", summary['traffic']['error_rate'] < 5 if summary['traffic']['total'] > 0 else True),
        ("Data Quality Acceptable", summary['tags']['by_quality']['bad'] < summary['tags']['total'] * 0.1 if summary['tags']['total'] > 0 else True),
        ("Response Time Acceptable", summary['traffic']['avg_rtt_ms'] < 500 if summary['traffic']['total'] > 0 else True),
    ]
    checklist_data = [["Check Item", "Status"]]
    for item, passed in checklist_items:
        checklist_data.append([item, "PASS" if passed else "FAIL"])
    
    checklist_table = Table(checklist_data, colWidths=[4*inch, 1.5*inch])
    checklist_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f3f4f6')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
    ]))
    # Color pass/fail cells
    for i, (_, passed) in enumerate(checklist_items, start=1):
        if passed:
            checklist_table.setStyle(TableStyle([
                ('BACKGROUND', (1, i), (1, i), colors.HexColor('#dcfce7')),
                ('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#166534')),
            ]))
        else:
            checklist_table.setStyle(TableStyle([
                ('BACKGROUND', (1, i), (1, i), colors.HexColor('#fee2e2')),
                ('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#991b1b')),
            ]))
    elements.append(checklist_table)
    
    # Footer
    elements.append(Spacer(1, 40))
    elements.append(Paragraph("This report was generated by ComGate - Tag Mapping Communication Gateway", 
                             ParagraphStyle('Footer', parent=normal_style, fontSize=8, textColor=colors.gray)))
    
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=FAT_Report_{project.get('name', project_id)}.pdf"}
    )

# ==================== DASHBOARD STATS ====================
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(
    project_id: Optional[str] = None,
    user: User = Depends(get_current_user)
):
    total_projects = await db.projects.count_documents({})
    
    # Scope devices/tags to project if selected
    device_filter = {"project_id": project_id} if project_id else {}
    tag_filter = {"project_id": project_id} if project_id else {}
    traffic_filter = {"project_id": project_id} if project_id else {}
    
    devices = await db.devices.count_documents(device_filter)
    tags = await db.tags.count_documents(tag_filter)
    online_filter = {**device_filter, "status": "online"}
    online_devices = await db.devices.count_documents(online_filter)
    
    # Gamification metrics
    first_project = await db.projects.find_one({}, sort=[("created_at", 1)])
    if first_project and "created_at" in first_project:
        try:
            if isinstance(first_project["created_at"], str):
                start_time = datetime.fromisoformat(first_project["created_at"].replace('Z', '+00:00'))
            else:
                start_time = first_project["created_at"]
            uptime_seconds = (datetime.now(timezone.utc) - start_time).total_seconds()
        except:
            uptime_seconds = 0
    else:
        uptime_seconds = 0
    
    # Calculate data quality score (percentage of tags with valid values)
    tags_with_values = await db.tags.count_documents({**tag_filter, "current_value": {"$ne": None}})
    data_quality_score = round((tags_with_values / tags * 100) if tags > 0 else 100, 1)
    
    # Calculate communication health (based on recent traffic)
    time_filter = {"timestamp": {"$gte": (datetime.now(timezone.utc) - timedelta(hours=1)).isoformat()}}
    recent_traffic = await db.traffic_logs.count_documents({**traffic_filter, **time_filter})
    
    # Count active protocols
    active_modbus_servers = len([s for s in modbus_servers.values() if s.is_running])
    active_modbus_clients = len([c for c in modbus_clients.values() if c.is_connected])
    active_opcua_servers = len([s for s in opcua_servers.values() if s.is_running])
    active_opcua_clients = len([c for c in opcua_clients.values() if c.is_connected])
    active_opcda_servers = len([s for s in opcda_servers.values() if s.is_running])
    active_opcda_clients = len([c for c in opcda_clients.values() if c.is_connected])
    
    # Calculate overall health score (0-100)
    health_factors = []
    if devices > 0:
        health_factors.append((online_devices / devices) * 100)
    health_factors.append(data_quality_score)
    if recent_traffic > 0:
        health_factors.append(min(100, recent_traffic / 10 * 100))
    overall_health = round(sum(health_factors) / len(health_factors) if health_factors else 100, 1)
    
    return {
        "projects": total_projects,
        "devices": devices,
        "tags": tags,
        "online_devices": online_devices,
        "active_simulators": len([s for s in simulator_sessions.values() if s.is_running]),
        "uptime_seconds": int(uptime_seconds),
        "uptime_formatted": format_uptime(uptime_seconds),
        "data_quality_score": data_quality_score,
        "overall_health": overall_health,
        "recent_traffic_count": recent_traffic,
        "tags_with_values": tags_with_values,
        "active_protocols": {
            "modbus_servers": active_modbus_servers,
            "modbus_clients": active_modbus_clients,
            "opcua_servers": active_opcua_servers,
            "opcua_clients": active_opcua_clients,
            "opcda_servers": active_opcda_servers,
            "opcda_clients": active_opcda_clients,
            "total_active": active_modbus_servers + active_modbus_clients + active_opcua_servers + active_opcua_clients + active_opcda_servers + active_opcda_clients
        }
    }

def format_uptime(seconds: float) -> str:
    """Format uptime in human-readable form"""
    if seconds < 60:
        return f"{int(seconds)}s"
    elif seconds < 3600:
        return f"{int(seconds // 60)}m {int(seconds % 60)}s"
    elif seconds < 86400:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        return f"{hours}h {minutes}m"
    else:
        days = int(seconds // 86400)
        hours = int((seconds % 86400) // 3600)
        return f"{days}d {hours}h"

# ==================== ANALYTICS APIs ====================

@api_router.get("/analytics/communication-stats")
async def get_communication_stats(
    time_range: str = "24h",  # 1h, 24h, 7d, 30d
    user: User = Depends(get_current_user)
):
    """Get communication statistics for analytics dashboard"""
    # Parse time range
    range_hours = {"1h": 1, "24h": 24, "7d": 168, "30d": 720}.get(time_range, 24)
    start_time = datetime.now(timezone.utc) - timedelta(hours=range_hours)
    
    # Get traffic logs for the time range
    traffic_query = {"timestamp": {"$gte": start_time.isoformat()}}
    total_requests = await db.traffic_logs.count_documents(traffic_query)
    
    success_query = {**traffic_query, "status": "ok"}
    success_count = await db.traffic_logs.count_documents(success_query)
    
    error_query = {**traffic_query, "status": {"$in": ["error", "timeout"]}}
    error_count = await db.traffic_logs.count_documents(error_query)
    
    # Calculate success rate
    success_rate = round((success_count / total_requests * 100) if total_requests > 0 else 100, 1)
    
    # Get average latency from traffic logs
    pipeline = [
        {"$match": traffic_query},
        {"$group": {
            "_id": None,
            "avg_latency": {"$avg": "$rtt_ms"},
            "max_latency": {"$max": "$rtt_ms"},
            "min_latency": {"$min": "$rtt_ms"}
        }}
    ]
    latency_stats = await db.traffic_logs.aggregate(pipeline).to_list(1)
    
    avg_latency = round(latency_stats[0]["avg_latency"], 2) if latency_stats and latency_stats[0].get("avg_latency") else 0
    max_latency = round(latency_stats[0]["max_latency"], 2) if latency_stats and latency_stats[0].get("max_latency") else 0
    min_latency = round(latency_stats[0]["min_latency"], 2) if latency_stats and latency_stats[0].get("min_latency") else 0
    
    # Get traffic by protocol
    protocol_pipeline = [
        {"$match": traffic_query},
        {"$group": {"_id": "$protocol", "count": {"$sum": 1}}}
    ]
    protocol_stats = await db.traffic_logs.aggregate(protocol_pipeline).to_list(10)
    traffic_by_protocol = {p["_id"]: p["count"] for p in protocol_stats if p["_id"]}
    
    # Get error breakdown
    error_pipeline = [
        {"$match": error_query},
        {"$group": {"_id": "$error_type", "count": {"$sum": 1}}}
    ]
    error_stats = await db.traffic_logs.aggregate(error_pipeline).to_list(10)
    errors_by_type = {e["_id"] or "unknown": e["count"] for e in error_stats}
    
    # Calculate throughput (requests per minute)
    throughput = round(total_requests / (range_hours * 60), 2) if range_hours > 0 else 0
    
    return {
        "time_range": time_range,
        "total_requests": total_requests,
        "success_count": success_count,
        "error_count": error_count,
        "success_rate": success_rate,
        "latency": {
            "avg_ms": avg_latency,
            "max_ms": max_latency,
            "min_ms": min_latency
        },
        "throughput_per_minute": throughput,
        "traffic_by_protocol": traffic_by_protocol,
        "errors_by_type": errors_by_type
    }

@api_router.get("/analytics/tag-health")
async def get_tag_health(
    project_id: Optional[str] = None,
    user: User = Depends(get_current_user)
):
    """Get tag health statistics"""
    query = {"project_id": project_id} if project_id else {}
    
    total_tags = await db.tags.count_documents(query)
    
    # Quality breakdown
    good_tags = await db.tags.count_documents({**query, "quality": "good"})
    bad_tags = await db.tags.count_documents({**query, "quality": "bad"})
    uncertain_tags = await db.tags.count_documents({**query, "quality": "uncertain"})
    
    # Tags with values
    tags_with_value = await db.tags.count_documents({**query, "current_value": {"$ne": None}})
    
    # Stale tags (not updated in last hour)
    stale_threshold = (datetime.now(timezone.utc) - timedelta(hours=1)).isoformat()
    stale_tags = await db.tags.count_documents({
        **query,
        "$or": [
            {"last_update": {"$lt": stale_threshold}},
            {"last_update": None}
        ]
    })
    
    # Permission breakdown
    read_only = await db.tags.count_documents({**query, "permission": "R"})
    write_only = await db.tags.count_documents({**query, "permission": "W"})
    read_write = await db.tags.count_documents({**query, "permission": "RW"})
    
    # Object type breakdown
    type_pipeline = [
        {"$match": query} if query else {"$match": {}},
        {"$group": {"_id": "$object_type", "count": {"$sum": 1}}}
    ]
    type_stats = await db.tags.aggregate(type_pipeline).to_list(10)
    tags_by_type = {t["_id"]: t["count"] for t in type_stats if t["_id"]}
    
    # Data type breakdown
    dtype_pipeline = [
        {"$match": query} if query else {"$match": {}},
        {"$group": {"_id": "$data_type", "count": {"$sum": 1}}}
    ]
    dtype_stats = await db.tags.aggregate(dtype_pipeline).to_list(20)
    tags_by_data_type = {d["_id"]: d["count"] for d in dtype_stats if d["_id"]}
    
    # Calculate health score
    health_score = 0
    if total_tags > 0:
        quality_score = (good_tags / total_tags) * 40
        value_score = (tags_with_value / total_tags) * 30
        freshness_score = ((total_tags - stale_tags) / total_tags) * 30
        health_score = round(quality_score + value_score + freshness_score, 1)
    
    return {
        "total_tags": total_tags,
        "health_score": health_score,
        "quality": {
            "good": good_tags,
            "bad": bad_tags,
            "uncertain": uncertain_tags
        },
        "values": {
            "with_value": tags_with_value,
            "without_value": total_tags - tags_with_value,
            "stale": stale_tags
        },
        "permissions": {
            "read_only": read_only,
            "write_only": write_only,
            "read_write": read_write
        },
        "by_object_type": tags_by_type,
        "by_data_type": tags_by_data_type
    }

@api_router.get("/analytics/protocol-comparison")
async def get_protocol_comparison(user: User = Depends(get_current_user)):
    """Compare performance across protocols"""
    
    protocols = ["modbus_tcp", "modbus_udp", "opcua", "opcda"]
    comparison = {}
    
    for protocol in protocols:
        traffic_query = {"protocol": protocol}
        total = await db.traffic_logs.count_documents(traffic_query)
        success = await db.traffic_logs.count_documents({**traffic_query, "status": "ok"})
        
        # Get latency stats
        pipeline = [
            {"$match": traffic_query},
            {"$group": {
                "_id": None,
                "avg_latency": {"$avg": "$rtt_ms"},
                "total_bytes": {"$sum": "$bytes_transferred"}
            }}
        ]
        stats = await db.traffic_logs.aggregate(pipeline).to_list(1)
        
        comparison[protocol] = {
            "total_requests": total,
            "success_count": success,
            "success_rate": round((success / total * 100) if total > 0 else 0, 1),
            "avg_latency_ms": round(stats[0]["avg_latency"], 2) if stats and stats[0].get("avg_latency") else 0,
            "total_bytes": stats[0].get("total_bytes", 0) if stats else 0
        }
    
    # Active connections
    comparison["active_connections"] = {
        "modbus_servers": len([s for s in modbus_servers.values() if s.is_running]),
        "modbus_clients": len([c for c in modbus_clients.values() if c.is_connected]),
        "opcua_servers": len([s for s in opcua_servers.values() if s.is_running]),
        "opcua_clients": len([c for c in opcua_clients.values() if c.is_connected]),
        "opcda_servers": len([s for s in opcda_servers.values() if s.is_running]),
        "opcda_clients": len([c for c in opcda_clients.values() if c.is_connected])
    }
    
    return comparison

@api_router.get("/analytics/time-series")
async def get_time_series_data(
    metric: str = "traffic",  # traffic, latency, errors, quality
    time_range: str = "24h",
    interval: str = "1h",  # 5m, 15m, 1h, 6h, 1d
    user: User = Depends(get_current_user)
):
    """Get time-series data for charts"""
    # Parse time range
    range_hours = {"1h": 1, "24h": 24, "7d": 168, "30d": 720}.get(time_range, 24)
    interval_minutes = {"5m": 5, "15m": 15, "1h": 60, "6h": 360, "1d": 1440}.get(interval, 60)
    
    start_time = datetime.now(timezone.utc) - timedelta(hours=range_hours)
    
    # Generate time buckets
    num_buckets = min(100, int(range_hours * 60 / interval_minutes))
    data_points = []
    
    for i in range(num_buckets):
        bucket_start = start_time + timedelta(minutes=i * interval_minutes)
        bucket_end = bucket_start + timedelta(minutes=interval_minutes)
        
        bucket_query = {
            "timestamp": {
                "$gte": bucket_start.isoformat(),
                "$lt": bucket_end.isoformat()
            }
        }
        
        if metric == "traffic":
            count = await db.traffic_logs.count_documents(bucket_query)
            data_points.append({
                "timestamp": bucket_start.isoformat(),
                "value": count
            })
        elif metric == "latency":
            pipeline = [
                {"$match": bucket_query},
                {"$group": {"_id": None, "avg": {"$avg": "$rtt_ms"}}}
            ]
            result = await db.traffic_logs.aggregate(pipeline).to_list(1)
            avg_latency = round(result[0]["avg"], 2) if result and result[0].get("avg") else 0
            data_points.append({
                "timestamp": bucket_start.isoformat(),
                "value": avg_latency
            })
        elif metric == "errors":
            error_count = await db.traffic_logs.count_documents({
                **bucket_query,
                "status": {"$in": ["error", "timeout"]}
            })
            data_points.append({
                "timestamp": bucket_start.isoformat(),
                "value": error_count
            })
        elif metric == "quality":
            # Simulate quality score trend (would need historical quality snapshots in production)
            good = await db.tags.count_documents({"quality": "good"})
            total = await db.tags.count_documents({})
            score = round((good / total * 100) if total > 0 else 100, 1)
            data_points.append({
                "timestamp": bucket_start.isoformat(),
                "value": score
            })
    
    return {
        "metric": metric,
        "time_range": time_range,
        "interval": interval,
        "data": data_points
    }

@api_router.get("/analytics/summary")
async def get_analytics_summary(user: User = Depends(get_current_user)):
    """Get a summary of all analytics for dashboard widgets"""
    # Communication stats (last 24h)
    comm_stats = await get_communication_stats("24h", user)
    
    # Tag health
    tag_health = await get_tag_health(None, user)
    
    # Active protocols
    active_protocols = {
        "modbus": len([s for s in modbus_servers.values() if s.is_running]) + len([c for c in modbus_clients.values() if c.is_connected]),
        "opcua": len([s for s in opcua_servers.values() if s.is_running]) + len([c for c in opcua_clients.values() if c.is_connected]),
        "opcda": len([s for s in opcda_servers.values() if s.is_running]) + len([c for c in opcda_clients.values() if c.is_connected])
    }
    
    # Peak hours analysis (simulated - would need historical data)
    peak_hour = 14  # 2 PM
    
    return {
        "communication": {
            "success_rate": comm_stats["success_rate"],
            "avg_latency_ms": comm_stats["latency"]["avg_ms"],
            "throughput": comm_stats["throughput_per_minute"],
            "total_requests_24h": comm_stats["total_requests"]
        },
        "tag_health": {
            "score": tag_health["health_score"],
            "good_percentage": round((tag_health["quality"]["good"] / tag_health["total_tags"] * 100) if tag_health["total_tags"] > 0 else 100, 1),
            "stale_count": tag_health["values"]["stale"]
        },
        "protocols": {
            "active_count": sum(active_protocols.values()),
            "breakdown": active_protocols
        },
        "insights": {
            "peak_hour": peak_hour,
            "recommendation": "OPC UA" if active_protocols.get("opcua", 0) == 0 else None
        }
    }

# ==================== HISTORIAN / TREND DATA ====================
class HistoricalDataRequest(BaseModel):
    tag_ids: List[str]
    start_time: datetime
    end_time: datetime
    interval_ms: int = 1000

class PredictionRequest(BaseModel):
    tag_id: str
    historical_values: List[float]
    steps: int = 20

@api_router.post("/historian/historical-data")
async def get_historical_data(
    request: HistoricalDataRequest,
    user: User = Depends(get_current_user)
):
    """
    Retrieve historical data for tags within a time range.
    Since we don't have a dedicated historian database, this generates simulated historical data
    based on the current tag values and random walk patterns.
    """
    result = {}
    
    for tag_id in request.tag_ids:
        tag = await db.tags.find_one({"id": tag_id}, {"_id": 0})
        if not tag:
            continue
            
        # Generate historical data points
        start_ts = request.start_time.timestamp() * 1000
        end_ts = request.end_time.timestamp() * 1000
        interval = request.interval_ms
        
        points = []
        num_points = min(500, int((end_ts - start_ts) / interval))
        
        # Start from current value or random
        current_val = tag.get('current_value')
        if current_val is None or not isinstance(current_val, (int, float)):
            current_val = 50.0
        
        value = float(current_val)
        import random
        
        for i in range(num_points):
            timestamp = start_ts + (i * interval)
            # Random walk with mean reversion
            value += (random.random() - 0.5) * 5
            value += (current_val - value) * 0.01  # Mean reversion
            
            # Apply min/max bounds if defined
            min_val = tag.get('min_value', 0)
            max_val = tag.get('max_value', 100)
            if min_val is not None:
                value = max(min_val, value)
            if max_val is not None:
                value = min(max_val, value)
            
            points.append({
                "time": int(timestamp),
                "value": round(value, 2),
                "quality": "good" if random.random() > 0.02 else "bad"
            })
        
        result[tag_id] = points
    
    return result

@api_router.post("/historian/predict")
async def predict_arima(
    request: PredictionRequest,
    user: User = Depends(get_current_user)
):
    """
    Generate ARIMA-based predictions for a tag.
    Uses statsmodels ARIMA model for time series forecasting.
    """
    try:
        import numpy as np
        from statsmodels.tsa.arima.model import ARIMA
        import warnings
        warnings.filterwarnings('ignore')
        
        values = np.array(request.historical_values)
        
        if len(values) < 10:
            return {"error": "Need at least 10 historical values for prediction"}
        
        # Fit ARIMA model (p=2, d=1, q=2 is a reasonable default)
        try:
            model = ARIMA(values, order=(2, 1, 2))
            fitted = model.fit()
            
            # Forecast
            forecast = fitted.forecast(steps=request.steps)
            
            # Calculate confidence intervals
            conf_int = fitted.get_forecast(steps=request.steps).conf_int()
            
            predictions = []
            for i in range(request.steps):
                predictions.append({
                    "step": i + 1,
                    "value": round(float(forecast.iloc[i] if hasattr(forecast, 'iloc') else forecast[i]), 2),
                    "lower_bound": round(float(conf_int.iloc[i, 0] if hasattr(conf_int, 'iloc') else conf_int[i, 0]), 2),
                    "upper_bound": round(float(conf_int.iloc[i, 1] if hasattr(conf_int, 'iloc') else conf_int[i, 1]), 2)
                })
            
            return {
                "tag_id": request.tag_id,
                "predictions": predictions,
                "model_info": {
                    "order": (2, 1, 2),
                    "aic": round(fitted.aic, 2) if hasattr(fitted, 'aic') else None
                }
            }
            
        except Exception as arima_error:
            # Fallback to simple linear extrapolation
            logger.warning(f"ARIMA failed, using linear extrapolation: {arima_error}")
            
            # Simple linear trend
            x = np.arange(len(values))
            slope = np.polyfit(x, values, 1)[0]
            last_value = values[-1]
            
            predictions = []
            for i in range(request.steps):
                pred_value = last_value + slope * (i + 1)
                predictions.append({
                    "step": i + 1,
                    "value": round(float(pred_value), 2),
                    "lower_bound": round(float(pred_value - 5), 2),
                    "upper_bound": round(float(pred_value + 5), 2)
                })
            
            return {
                "tag_id": request.tag_id,
                "predictions": predictions,
                "model_info": {"method": "linear_extrapolation"}
            }
            
    except Exception as e:
        logger.error(f"Prediction error: {e}")
        raise HTTPException(status_code=500, detail=f"Prediction failed: {str(e)}")

@api_router.get("/historian/tag-values/{project_id}")
async def get_live_tag_values(
    project_id: str,
    tag_ids: str = Query(..., description="Comma-separated tag IDs"),
    user: User = Depends(get_current_user)
):
    """Get current values for multiple tags (for real-time trend updates)"""
    ids = [t.strip() for t in tag_ids.split(',') if t.strip()]
    
    tags = await db.tags.find(
        {"id": {"$in": ids}, "project_id": project_id},
        {"_id": 0, "id": 1, "name": 1, "current_value": 1, "quality": 1, "last_update": 1, "alarm_enable": 1}
    ).to_list(100)
    
    return {tag['id']: tag for tag in tags}

# ==================== WEBSOCKET FOR REAL-TIME UPDATES ====================
class ConnectionManager:
    """Manages WebSocket connections for real-time tag updates"""
    
    def __init__(self):
        self.active_connections: Dict[str, List[WebSocket]] = {}  # project_id -> connections
        self.all_connections: List[WebSocket] = []
    
    async def connect(self, websocket: WebSocket, project_id: Optional[str] = None):
        await websocket.accept()
        self.all_connections.append(websocket)
        if project_id:
            if project_id not in self.active_connections:
                self.active_connections[project_id] = []
            self.active_connections[project_id].append(websocket)
    
    def disconnect(self, websocket: WebSocket, project_id: Optional[str] = None):
        if websocket in self.all_connections:
            self.all_connections.remove(websocket)
        if project_id and project_id in self.active_connections:
            if websocket in self.active_connections[project_id]:
                self.active_connections[project_id].remove(websocket)
    
    async def broadcast_to_project(self, project_id: str, message: dict):
        """Broadcast message to all connections watching a project"""
        if project_id in self.active_connections:
            for connection in self.active_connections[project_id]:
                try:
                    await connection.send_json(message)
                except:
                    pass
    
    async def broadcast_all(self, message: dict):
        """Broadcast message to all connections"""
        for connection in self.all_connections:
            try:
                await connection.send_json(message)
            except:
                pass

ws_manager = ConnectionManager()

@app.websocket("/ws/tags/{project_id}")
async def websocket_tags(websocket: WebSocket, project_id: str):
    """WebSocket endpoint for real-time tag value updates"""
    await ws_manager.connect(websocket, project_id)
    try:
        while True:
            # Wait for any message (ping/pong or subscription changes)
            data = await websocket.receive_text()
            
            if data == "ping":
                await websocket.send_text("pong")
            elif data.startswith("subscribe:"):
                # Client can subscribe to specific tags
                tag_ids = data.replace("subscribe:", "").split(",")
                await websocket.send_json({
                    "type": "subscribed",
                    "tags": tag_ids
                })
                
    except WebSocketDisconnect:
        ws_manager.disconnect(websocket, project_id)

@app.websocket("/ws/system")
async def websocket_system(websocket: WebSocket):
    """WebSocket endpoint for system-wide updates (protocol status, etc.)"""
    await ws_manager.connect(websocket)
    try:
        while True:
            data = await websocket.receive_text()
            
            if data == "ping":
                await websocket.send_text("pong")
            elif data == "status":
                # Send current system status
                status = {
                    "type": "system_status",
                    "modbus_servers": len([s for s in modbus_servers.values() if s.is_running]),
                    "modbus_clients": len([c for c in modbus_clients.values() if c.is_connected]),
                    "opcua_servers": len([s for s in opcua_servers.values() if s.is_running]),
                    "opcua_clients": len([c for c in opcua_clients.values() if c.is_connected]),
                    "opcda_servers": len([s for s in opcda_servers.values() if s.is_running]),
                    "opcda_clients": len([c for c in opcda_clients.values() if c.is_connected]),
                    "timestamp": datetime.now(timezone.utc).isoformat()
                }
                await websocket.send_json(status)
                
    except WebSocketDisconnect:
        ws_manager.disconnect(websocket)

# Helper function to broadcast tag updates (call this when tags are updated)
async def broadcast_tag_update(project_id: str, tag_id: str, value: Any, quality: str = "good"):
    """Broadcast a tag value update to all connected clients"""
    await ws_manager.broadcast_to_project(project_id, {
        "type": "tag_update",
        "tag_id": tag_id,
        "value": value,
        "quality": quality,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })

# ==================== HEALTH CHECK ====================
@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# ==================== NETWORK DISCOVERY ====================
import socket

def get_local_ip():
    """Get local IP address"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

@api_router.get("/discovery/announce")
async def discovery_announce():
    """Announce this server's presence for network discovery"""
    local_ip = get_local_ip()
    active_modbus = len([s for s in modbus_servers.values() if s.is_running])
    active_opcua = len([s for s in opcua_servers.values() if s.is_running])
    return {
        "service": "ComGate",
        "ip": local_ip,
        "port": 8001,
        "modbus_servers_active": active_modbus,
        "opcua_servers_active": active_opcua,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

class NetworkScanRequest(BaseModel):
    subnet: Optional[str] = None
    timeout: float = 0.5

class ModbusScanRequest(BaseModel):
    subnet: Optional[str] = None
    port: int = 5020
    timeout: float = 0.5
    probe_modbus: bool = True  # Actually verify Modbus protocol
    scan_ports: List[int] = [502, 503, 5020]  # Common Modbus ports

class AddDiscoveredDeviceRequest(BaseModel):
    project_id: str
    ip: str
    port: int
    device_name: Optional[str] = None
    unit_id: int = 1

@api_router.post("/discovery/scan-modbus")
async def discovery_scan_modbus(request: ModbusScanRequest, user: User = Depends(get_current_user)):
    """Scan for Modbus TCP devices on the network with protocol verification"""
    import socket
    
    local_ip = get_local_ip()
    subnet_prefix = request.subnet or ".".join(local_ip.split(".")[:3])
    scan_ports = request.scan_ports if request.scan_ports else [request.port]
    scan_timeout = request.timeout
    probe_modbus = request.probe_modbus
    
    found_devices = []
    
    async def check_modbus_host(ip, port):
        def _check():
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(scan_timeout)
                result = sock.connect_ex((ip, port))
                
                if result == 0:
                    device_info = {
                        "ip": ip,
                        "port": port,
                        "type": "modbus_tcp",
                        "is_self": (ip == local_ip),
                        "verified": False,
                        "unit_ids": [],
                        "hostname": None
                    }
                    
                    # Try to get hostname
                    try:
                        hostname = socket.gethostbyaddr(ip)[0]
                        device_info["hostname"] = hostname
                    except:
                        pass
                    
                    # Probe Modbus if enabled
                    if probe_modbus:
                        try:
                            # Send Modbus TCP request: Read Holding Register (FC 03) at address 0, quantity 1
                            # Transaction ID (2) + Protocol ID (2) + Length (2) + Unit ID (1) + FC (1) + Start (2) + Qty (2)
                            for unit_id in [1, 0, 255]:  # Try common unit IDs
                                try:
                                    modbus_req = bytes([
                                        0x00, 0x01,  # Transaction ID
                                        0x00, 0x00,  # Protocol ID (Modbus)
                                        0x00, 0x06,  # Length
                                        unit_id,     # Unit ID
                                        0x03,        # Function Code: Read Holding Registers
                                        0x00, 0x00,  # Start Address
                                        0x00, 0x01   # Quantity
                                    ])
                                    
                                    probe_sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                                    probe_sock.settimeout(1.0)
                                    probe_sock.connect((ip, port))
                                    probe_sock.send(modbus_req)
                                    response = probe_sock.recv(256)
                                    probe_sock.close()
                                    
                                    # Check if valid Modbus response (at least 9 bytes, protocol ID = 0)
                                    if len(response) >= 9 and response[2:4] == b'\x00\x00':
                                        device_info["verified"] = True
                                        if unit_id not in device_info["unit_ids"]:
                                            device_info["unit_ids"].append(unit_id)
                                except:
                                    pass
                        except:
                            pass
                    
                    sock.close()
                    return device_info
                
                sock.close()
            except Exception:
                pass
            return None
        
        result = await asyncio.get_event_loop().run_in_executor(None, _check)
        if result:
            found_devices.append(result)
    
    # Scan all IPs on all specified ports
    tasks = []
    for port in scan_ports:
        for i in range(1, 255):
            tasks.append(check_modbus_host(f"{subnet_prefix}.{i}", port))
    
    await asyncio.gather(*tasks)
    
    # Deduplicate by IP (keep the one with most info)
    devices_by_ip = {}
    for d in found_devices:
        key = d["ip"]
        if key not in devices_by_ip:
            devices_by_ip[key] = d
        else:
            # Merge unit_ids and prefer verified
            existing = devices_by_ip[key]
            if d.get("verified"):
                existing["verified"] = True
            existing["unit_ids"] = list(set(existing.get("unit_ids", []) + d.get("unit_ids", [])))
    
    unique_devices = list(devices_by_ip.values())
    unique_devices.sort(key=lambda x: (not x.get("verified", False), not x.get("is_self", False)))
    
    return {
        "local_ip": local_ip,
        "subnet": subnet_prefix,
        "ports_scanned": scan_ports,
        "port_scanned": scan_ports[0],  # For backward compatibility
        "devices": unique_devices,
        "total_found": len(unique_devices),
        "verified_count": sum(1 for d in unique_devices if d.get("verified")),
        "scanned_at": datetime.now(timezone.utc).isoformat()
    }


@api_router.post("/discovery/add-device")
async def discovery_add_device(request: AddDiscoveredDeviceRequest, user: User = Depends(require_role([UserRole.ADMIN, UserRole.ENGINEER]))):
    """Add a discovered Modbus device to a project"""
    
    # Check project exists
    project = await db.projects.find_one({"id": request.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Generate device name if not provided
    device_name = request.device_name or f"Device_{request.ip.replace('.', '_')}_{request.port}"
    
    # Check if device already exists
    existing = await db.devices.find_one({
        "project_id": request.project_id,
        "ip_address": request.ip,
        "port": request.port
    }, {"_id": 0})
    
    if existing:
        return {"message": "Device already exists", "device": existing, "created": False}
    
    # Create device
    device = Device(
        name=device_name,
        project_id=request.project_id,
        protocol=ProtocolType.TCP,
        ip_address=request.ip,
        port=request.port,
        unit_id=request.unit_id,
        timeout_ms=3000,
        retries=3,
        max_block_size=120,
        default_endian="ABCD"
    )
    
    doc = device.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.devices.insert_one(doc)
    
    # Log audit
    await log_audit(user, "device_added_from_discovery", {
        "device_id": device.id,
        "device_name": device_name,
        "ip": request.ip,
        "port": request.port
    })
    
    return {"message": "Device added successfully", "device": doc, "created": True}

@api_router.post("/discovery/scan")
async def discovery_scan(request: NetworkScanRequest, user: User = Depends(get_current_user)):
    """Scan the local network for ComGate servers using stdlib only"""
    import urllib.request
    
    local_ip = get_local_ip()
    subnet_prefix = request.subnet or ".".join(local_ip.split(".")[:3])
    scan_timeout = request.timeout
    
    found_servers = []
    
    async def check_host(ip):
        def _check():
            try:
                url = f"http://{ip}:8001/api/discovery/announce"
                req = urllib.request.Request(url, headers={"Accept": "application/json"})
                with urllib.request.urlopen(req, timeout=scan_timeout) as resp:
                    if resp.status == 200:
                        data = json.loads(resp.read().decode())
                        if data.get("service") == "ComGate":
                            data["is_self"] = (ip == local_ip)
                            return data
            except Exception:
                pass
            return None
        
        result = await asyncio.get_event_loop().run_in_executor(None, _check)
        if result:
            found_servers.append(result)
    
    # Scan IPs 1-254 in the subnet concurrently
    tasks = [check_host(f"{subnet_prefix}.{i}") for i in range(1, 255)]
    await asyncio.gather(*tasks)
    
    found_servers.sort(key=lambda x: x.get("is_self", False), reverse=True)
    return {
        "local_ip": local_ip,
        "subnet": subnet_prefix,
        "servers": found_servers,
        "scanned_at": datetime.now(timezone.utc).isoformat()
    }

# Download endpoint for deployment package
@api_router.get("/download/comgate-package")
async def download_comgate_package():
    """Download the complete ComGate deployment package"""
    zip_path = Path(__file__).parent.parent / "comgate-complete.zip"
    if not zip_path.exists():
        raise HTTPException(status_code=404, detail="Package not found")
    return FileResponse(
        path=str(zip_path),
        filename="comgate-complete.zip",
        media_type="application/zip"
    )

# Include the router in the main app
app.include_router(api_router)

# Serve pre-built frontend (for local deployment without Node.js)
frontend_build = Path(__file__).parent.parent / "frontend" / "build"
if frontend_build.exists():
    app.mount("/static", StaticFiles(directory=str(frontend_build / "static")), name="static")
    
    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        # Serve actual files if they exist
        file_path = frontend_build / full_path
        if full_path and file_path.exists() and file_path.is_file():
            return FileResponse(str(file_path))
        # Otherwise serve index.html (React router handles the rest)
        return FileResponse(str(frontend_build / "index.html"))

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup():
    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)
    await db.projects.create_index("id", unique=True)
    await db.devices.create_index("id", unique=True)
    await db.devices.create_index("project_id")
    await db.tags.create_index("id", unique=True)
    await db.tags.create_index("project_id")
    await db.tags.create_index("device_id")
    await db.tags.create_index("name")
    await db.traffic_logs.create_index("project_id")
    await db.traffic_logs.create_index("timestamp")
    await db.audit_logs.create_index("timestamp")
    
    # Create default admin user if not exists
    admin = await db.users.find_one({"email": "admin@comgate.com"}, {"_id": 0})
    if not admin:
        admin_user = User(
            email="admin@comgate.com",
            username="admin",
            role=UserRole.ADMIN
        )
        doc = admin_user.model_dump()
        doc['password_hash'] = hash_password("admin123")
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        logger.info("Created default admin user: admin@comgate.com / admin123")

@app.on_event("shutdown")
async def shutdown_db_client():
    # Stop all polling tasks
    for engine in polling_tasks.values():
        await engine.stop()
    # Stop all Modbus servers
    for server in list(modbus_servers.values()):
        await server.stop()
    # Stop all simulations
    for sim in list(simulation_tasks.values()):
        await sim.stop()
    # Stop all Modbus clients
    for mb_client in list(modbus_clients.values()):
        await mb_client.stop_polling()
    # Stop all OPC UA servers
    for opcua_srv in list(opcua_servers.values()):
        await opcua_srv.stop()
    # Stop all OPC UA clients
    for opcua_cli in list(opcua_clients.values()):
        await opcua_cli.stop_polling()
    # Stop all OPC DA servers
    for opcda_srv in list(opcda_servers.values()):
        await opcda_srv.stop()
    # Stop all OPC DA clients
    for opcda_cli in list(opcda_clients.values()):
        await opcda_cli.stop_polling()
    # Close MongoDB client
    client.close()
